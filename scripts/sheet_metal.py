"""
Author: Manuel Rosa

TDLR: Automatically extracts DXFs analyses them and creates multiple tables/images with information

Description: 
- Analyses all dxfs and atributes them different types of cuts
- Nesting of all sheets
- Creates list of all DXF
- Creates a list of the necessary sheets for the project
- Creates Images of all dxf
- Combines all images into 1 with legends
- Creates structured folders for every DXF file to ease the cutting of them
"""

import os
import pandas as pd
import re
import tkinter as tk
from tkinter import messagebox
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import shutil
import win32com.client as win32
import sys
import ezdxf
import math
import matplotlib.pyplot as plt
from ezdxf.addons.drawing import RenderContext, Frontend
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
from PIL import Image, ImageDraw, ImageFont
from PyPDF4.pdf import PdfFileReader, PdfFileWriter
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import io
import csv
import glob
from rectpack import newPacker
from sheet_phc_checker import find_code_sheet

print("A organizar dados...")


def convert_to_numeric(value):
    try:
        return float(value)
    except ValueError:
        return value


script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
parent_dir = os.path.dirname(script_dir)
source_name = os.path.join(parent_dir, "macros", "source_path.txt")
options = os.path.join(parent_dir, "extras", "op.txt")
with open(options, "r") as file:
    lines = file.readlines()
if lines:
    phc_data = lines[9].strip()
    if phc_data:
        phc_path = (((phc_data).split("="))[1])[1:]
csv_path = os.path.join(phc_path, "dxf_ana.csv")
excel_path = os.path.join(phc_path, "dxf_ana.xlsx")
excel_file = pd.read_excel(excel_path)
excel_file.to_csv(csv_path, index=False)

with open(source_name, "r") as file:
    source_path = file.read().strip()
parent_directory = os.path.dirname(source_path)
ass_name = os.path.basename(source_path)
ass_name = os.path.splitext(ass_name)[0]
folder_new = parent_directory + "//Listas_" + ass_name

with open(options, "r") as file:
    lines = file.readlines()
edit_line = lines[7].strip()
edit_val = edit_line[-1]
chapa_line = lines[4].strip()  
nesting = eval(lines[6].split("=")[1])
phc_location = lines[9].strip()
phc_location = ((phc_location.split("="))[1])[1:]
sheet_phc_location = os.path.join(phc_location, "Chapa")


coordinate_strings = re.findall(r"\((\d+), (\d+)\)", chapa_line)
n_plate = [(int(x), int(y)) for x, y in coordinate_strings]
plate_error = None

csv_data = {}
acronimo_mapping = {}


output_file = folder_new + "//" + ass_name + "_DXF.xlsx"
output_folder = folder_new + "//DXF"
output_pdf = folder_new + "//" + ass_name + "_DXF.pdf"
output_nesting_laser_pdf = folder_new + "//" + ass_name + "_Nesting.pdf"
worksheet_name = "DXF"
worksheet_name_l = "DXF_Nest"
logo_image_path = os.path.join(parent_dir, "extras", "Logo.pdf")
image_path = folder_new + "//Imagens"


if os.path.exists(output_file):
    try:
        print("being removed")
        os.remove(output_file)
        shutil.rmtree(output_folder)
        shutil.rmtree(image_path)
        os.remove(output_pdf)
        os.remove(output_nesting_laser_pdf)
    except:
        pass
if not os.path.exists(folder_new):
    os.makedirs(folder_new)


def showMessage(message, type="info", timeout=5000):
    import tkinter as tk
    from tkinter import messagebox as msgb

    root = tk.Tk()
    root.withdraw()
    try:
        root.after(timeout, root.destroy)
        if type == "info":
            msgb.showinfo("Terminado", message, master=root)
        elif type == "warning":
            msgb.showwarning("Warning", message, master=root)
        elif type == "error":
            msgb.showerror("Error", message, master=root)
    except:
        pass


def convert_worksheet_to_pdf(excel_file_path, worksheet_name, pdf_file_path):
    excel_app = win32.gencache.EnsureDispatch("Excel.Application")
    workbook = excel_app.Workbooks.Open(excel_file_path)
    worksheet = workbook.Worksheets(worksheet_name)

    columns_range = worksheet.Range("A:F")
    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = False

    pdf_export_format = win32.constants.xlTypePDF
    export_path = os.path.join(pdf_file_path)
    columns_range.ExportAsFixedFormat(pdf_export_format, export_path)

    workbook.Close(False)
    excel_app.Quit()


def put_watermark(input_pdf, output_pdf, watermark):
    watermark_instance = PdfFileReader(watermark)
    watermark_page = watermark_instance.getPage(0)

    scale_factor = 1 / 8  
    scaled_watermark_height = watermark_page.mediaBox.getUpperRight_y() * scale_factor
    watermark_page.scaleBy(scale_factor)
    pdf_reader = PdfFileReader(input_pdf)
    pdf_writer = PdfFileWriter()
    for page_num in range(pdf_reader.getNumPages()):
        page = pdf_reader.getPage(page_num)

        if page_num == 0:  
            page_width = page.mediaBox.getUpperRight_x()
            page_height = page.mediaBox.getUpperRight_y()
            page_width = float(page_width)
            page_height = float(page_height)
            x = 47 
            y = (
                page_height - scaled_watermark_height
            )  
            page.mergeTranslatedPage(watermark_page, x, y)
        pdf_writer.addPage(page)
    with open(output_pdf, "wb") as out:
        pdf_writer.write(out)


def create_page_pdf(num):
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)

    for i in range(1, num + 1):
        text = f"{i}/{num}"

        text_width = c.stringWidth(text, "Helvetica", 12)
        x = (210 - text_width) * mm  
        y = 4 * mm  

        c.drawString(x, y, text)
        c.showPage()

    c.save()

    packet.seek(0)
    return PdfFileReader(packet)

def add_page_numbers(pdf_path):
    """
    Add page numbers to a pdf, save the result as a new pdf
    @param pdf_path: path to pdf
    """
    output_pdf = PdfFileWriter()

    with open(pdf_path, "rb") as f:
        reader = PdfFileReader(f, strict=False)
        n = reader.getNumPages()
        number_pdf = create_page_pdf(n)
        for p in range(n):
            page = reader.getPage(p)
            number_layer_page = number_pdf.getPage(p)
            page.mergePage(number_layer_page)
            output_pdf.addPage(page)
        if output_pdf.getNumPages():
            newpath = pdf_path[:-4] + "_numbered.pdf"
            with open(newpath, "wb") as f:
                output_pdf.write(f)
        shutil.move(newpath, pdf_path)



def delete_files_with_pattern(folder_path, pattern):
    files_to_delete = glob.glob(os.path.join(folder_path, f"*{pattern}*"))
    for file_path in files_to_delete:
        os.remove(file_path)


def add_image_link(cell, exten_link, path_to_real):
    cell_value = cell.value
    if not cell_value or cell_value == "DXF":
        return

    image_link = os.path.relpath(
        os.path.join(image_folder, str(cell_value) + exten_link),
        os.path.dirname(path_to_real),
    )

    cell.hyperlink = image_link


def process_cut_list(cut_list):
    def reorder_tuple(tuple_item):
        if isinstance(tuple_item[0], tuple):
            first_values = tuple_item[0]
            return (max(first_values), min(first_values), tuple_item[1])
        return (max(tuple_item[:2]), min(tuple_item[:2]), tuple_item[2])


    leftover_plate = []
    matching_counters = {n_item: 0 for n_item in n_plate}

    cut_list = [(cut_value1, cut_value2) for cut_value1, cut_value2 in cut_list]

    plate_data = pd.DataFrame(
        columns=["cut", "Matched_Plate", "Original_Plate", "Match_Type"]
    )

    cut_list_error = []

    for cut_item in cut_list:
        cut_value1, cut_value2 = cut_item
        matched = 3
        closest_match = None
        min_distance = float("inf")

        for leftover_item in leftover_plate:
            leftover_value1, leftover_value2, leftover_value3 = leftover_item
            if leftover_value1 >= cut_value1 and leftover_value2 >= cut_value2:
                distance = abs(leftover_value1 - cut_value1) + abs(
                    leftover_value2 - cut_value2
                )
                if distance < min_distance:
                    closest_match = leftover_item
                    min_distance = distance
                    matched = 1
        if matched != 1:
            for n_item in n_plate:
                n_value1, n_value2 = n_item
                if n_value1 >= cut_value1 and n_value2 >= cut_value2:
                    distance = abs(n_value1 - cut_value1) + abs(n_value2 - cut_value2)
                    if distance < min_distance:
                        closest_match = n_item
                        min_distance = distance
                        matched = 2

        if matched == 1:
            closest_l_item = closest_match
            leftover_value1 = max(closest_l_item[0], 0)
            leftover_value2 = max(closest_l_item[1], 0)
            leftover_value3 = closest_l_item[2]
            leftover_plate.append(
                (leftover_value1, leftover_value2 - cut_value2, leftover_value3)
            )
            leftover_plate.append(
                (cut_value2, leftover_value1 - cut_value1, leftover_value3)
            )
            leftover_plate.remove(closest_l_item)
            leftover_plate = [reorder_tuple(item) for item in leftover_plate]
            leftover_plate = [
                item
                for item in leftover_plate
                if not any(
                    val == 0
                    for val in item
                    if isinstance(val, int) or isinstance(val, tuple)
                )
            ]
            plate_data = plate_data._append(
                {
                    "cut": f"{cut_value1}x{cut_value2}",
                    "Matched_Plate": closest_l_item[2],
                    "Original_Plate": closest_l_item[:2],
                    "Match_Type": "Leftover",
                },
                ignore_index=True,
            )
            leftover_plate = [reorder_tuple(item) for item in leftover_plate]
            leftover_plate = [
                item
                for item in leftover_plate
                if not any(
                    int(val) == 0
                    for val in item
                    if isinstance(val, int) or isinstance(val, tuple)
                )
            ]

        if matched == 2:
            closest_n_item = closest_match
            leftover_value1 = max(closest_n_item[0], 0)
            leftover_value2 = max(closest_n_item[1], 0)
            leftover_plate.append(
                (
                    leftover_value1,
                    leftover_value2 - cut_value2,
                    f"{closest_n_item}{matching_counters[closest_match]}",
                )
            )
            leftover_plate.append(
                (
                    cut_value2,
                    leftover_value1 - cut_value1,
                    f"{closest_n_item}{matching_counters[closest_match]}",
                )
            )
            matching_counters[closest_match] += 1
            leftover_plate = [reorder_tuple(item) for item in leftover_plate]
            leftover_plate = [
                item
                for item in leftover_plate
                if not any(
                    val == 0
                    for val in item
                    if isinstance(val, int) or isinstance(val, tuple)
                )
            ]
            plate_data = plate_data._append(
                {
                    "cut": f"{cut_value1}x{cut_value2}",
                    "Matched_Plate": f"{closest_n_item}{matching_counters[closest_match] - 1}",
                    "Original_Plate": closest_n_item[:2],
                    "Match_Type": "N_Type",
                },
                ignore_index=True,
            )
            leftover_plate = [reorder_tuple(item) for item in leftover_plate]
            leftover_plate = [
                item
                for item in leftover_plate
                if not any(
                    int(val) == 0
                    for val in item
                    if isinstance(val, int) or isinstance(val, tuple)
                )
            ]

        if matched == 3:
            cut_list_error.append(cut_item[:2])
    leftover_plate_org = []
    for item in leftover_plate:
        formatted_item = (
            f"{item[0]}x{item[1]}",
            item[2].replace("(", "").replace(", ", "x").replace(")", "_"),
        )
        leftover_plate_org.append(formatted_item)
    return plate_data, cut_list_error



def pack_rectangles(dictionary):
    def pack_with_bins(bins, rectangles):
        packer = newPacker()
        for r in rectangles:
            packer.add_rect(*r)
        for b in bins:
            packer.add_bin(*b)
        packer.pack()

        return packer.rect_list()

    results = [] 
    for key, value in dictionary.items():
        material, thickness = key.split("_")[1:3]
        b_sheet = value
        bins = []
        bin_loop = []
        plate_error = 0
        largest_plate = max(b_sheet, key=lambda x: x[0])
        largest_plate_2 = max(b_sheet, key=lambda x: x[1])
        largest_first_value = largest_plate[0]
        largest_second_value = largest_plate_2[1]
        larg_n_plate = max(n_plate, key=lambda x: x[0])
        n_plate_desc = sorted(n_plate, key=lambda x: x[0] + x[1])
        if (larg_n_plate[0] < largest_first_value) or (
            larg_n_plate[1] < largest_second_value
        ):
            plate_error = 1
        while b_sheet != [] and plate_error == 0:
            for plate in n_plate_desc:
                bin_loop.append(plate)
                tiny_rects = pack_with_bins(bin_loop, b_sheet)
                fitted_rects = [(x, y) for _, _, _, x, y, _ in tiny_rects]
                fixed_fitted_rects = [
                    (y, x) if x < y else (x, y) for x, y in fitted_rects
                ]
                bin_loop = []
                if len(tiny_rects) > 0:
                    for item in fixed_fitted_rects:
                        if item in b_sheet:
                            b_sheet.remove(item)
                    bins.append(plate)
                    break
        if plate_error == 0:
            bin_counts = {}
            for bin_size in bins:
                bin_str = f"{bin_size[0]}x{bin_size[1]}"
                bin_counts[bin_str] = bin_counts.get(bin_str, 0) + 1

            bin_str_list = []
            for bin_str, count in bin_counts.items():
                if count > 1:
                    bin_str_list.append(f"{bin_str}(x{count})")
                else:
                    bin_str_list.append(bin_str)
            results.append([material, thickness, ", ".join(bin_str_list)])
        else:
            results.append(
                [material, thickness, "".join(f"Erro Chapa: {largest_plate}")]
            )

    df = pd.DataFrame(results, columns=["Material", "Esp.", "Chapas"])
    return df

with open(csv_path, "r") as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        designacao = row["Designacao"]
        csv_data[designacao] = {
            "Acronimo": row["Acronimo"],
            "Esp. Min.": row["Esp. Min."],
            "Esp. Max.": row["Esp. Max."],
            "Min. n linha": row["Min. n linha"],
            "Max. n linha": row["Max. n linha"],
            "Min. dist. linha": row["Min. dist. linha"],
            "Max. dist. linha": row["Max. dist. linha"],
            "Min. n circ.": row["Min. n circ."],
            "Max. n circ.": row["Max. n circ."],
            "Min. dia. circ.": row["Min. dia. circ."],
            "Max. dia. circ.": row["Max. dia. circ."],
            "Min. n arc.": row["Min. n arc."],
            "Max. n arc.": row["Max. n arc."],
            "Min. dia. arc.": row["Min. dia. arc."],
            "Max. dia. arc.": row["Max. dia. arc."],
        }

designacao_list = acronimo_list = [
    csv_data[designacao]["Acronimo"] for designacao in csv_data
]

for designacao, data in csv_data.items():
    acronimo = data["Acronimo"]
    acronimo_mapping[acronimo] = designacao
folder_path = r"C:\ProgramData\DXFs_Macro"

if edit_val == "0":
    pattern_to_match = "_C."
    delete_files_with_pattern(folder_path, pattern_to_match)


def process_dxf_folder(folder_path):
    def process_dxf_file(input_file_path):
        doc = ezdxf.readfile(input_file_path)
        modelspace = doc.modelspace()
        entities = list(modelspace.query("*"))
        nested_entities = []
        num_lines = 0
        num_circles = 0
        num_arcs = 0
        different_handles = set()
        max_line_length = 0
        min_line_length = 0
        max_arc_diameter = 0
        min_arc_diameter = 0
        max_circ_diameter = 0
        min_circ_diameter = 0
        thick = float(re.search(r"_Thick\.(.*?)_", input_file_path).group(1))
        if entities:
            for entity in entities:
                entity_type = entity.dxftype()
                handle = entity.dxf.handle
                first_char = handle[0]

                if entity_type not in ("TEXT", "MTEXT", "ATTRIB", "ATTDEF"):
                    nested_entities.append(entity)

                if entity_type == "CIRCLE":
                    radius = entity.dxf.radius
                    diameter = 2 * radius
                    num_circles += 1
                    if diameter > max_circ_diameter:
                        max_circ_diameter = diameter
                    if diameter < min_circ_diameter:
                        min_circ_diameter = diameter

                if entity_type == "LINE":
                    num_lines += 1
                    start_point = entity.dxf.start
                    end_point = entity.dxf.end
                    dx = end_point[0] - start_point[0]
                    dy = end_point[1] - start_point[1]
                    length = math.sqrt(dx**2 + dy**2)
                    if length > max_line_length:
                        max_line_length = length
                    if length < min_line_length:
                        min_line_length = length

                if entity_type == "ARC":
                    radius = entity.dxf.radius
                    diameter = 2 * radius
                    num_arcs += 1
                    if diameter > max_arc_diameter:
                        max_arc_diameter = diameter
                    if diameter > min_arc_diameter:
                        min_arc_diameter = diameter
                if first_char not in different_handles:
                    different_handles.add(first_char)

            for designacao in csv_data:
                if csv_data[designacao]["Esp. Min."] != "":
                    if float(csv_data[designacao]["Esp. Min."]) > thick:
                        continue
                if csv_data[designacao]["Esp. Max."] != "":
                    if float(csv_data[designacao]["Esp. Max."]) < thick:
                        continue
                if csv_data[designacao]["Min. n linha"] != "":
                    if int(float(csv_data[designacao]["Min. n linha"])) > num_lines:
                        continue
                if csv_data[designacao]["Max. n linha"] != "":
                    if int(float(csv_data[designacao]["Max. n linha"])) < num_lines:
                        continue
                if csv_data[designacao]["Min. dist. linha"] != "":
                    if float(csv_data[designacao]["Min. dist linha"]) > min_line_length:
                        continue
                if csv_data[designacao]["Max. dist. linha"] != "":
                    if float(csv_data[designacao]["Max. dist linha"]) < max_line_length:
                        continue
                if csv_data[designacao]["Min. n circ."] != "":
                    if int(float(csv_data[designacao]["Min. n circ."])) > num_circles:
                        continue
                if csv_data[designacao]["Max. n circ."] != "":
                    if int(float(csv_data[designacao]["Max. n circ."])) < num_circles:
                        continue
                if csv_data[designacao]["Min. dia. circ."] != "":
                    if (
                        float(csv_data[designacao]["Min. dia. circ."])
                        > min_circ_diameter
                    ):
                        continue
                if csv_data[designacao]["Max. dia. circ."] != "":
                    if (
                        float(csv_data[designacao]["Max. dia. circ."])
                        < max_circ_diameter
                    ):
                        continue
                if csv_data[designacao]["Min. n arc."] != "":
                    if int(float(csv_data[designacao]["Min. n arc."])) > num_arcs:
                        continue
                if csv_data[designacao]["Max. n arc."] != "":
                    if int(float(csv_data[designacao]["Max. n arc."])) < num_arcs:
                        continue
                if csv_data[designacao]["Min. dia. arc."] != "":
                    if float(csv_data[designacao]["Min. dia. arc."]) > min_arc_diameter:
                        continue
                if csv_data[designacao]["Max. dia. arc."] != "":
                    if float(csv_data[designacao]["Max. dia. arc."]) < max_arc_diameter:
                        continue
                new_suffix = "_C." + csv_data[designacao]["Acronimo"]
                break
            if edit_val == "0":
                filename, _ = os.path.splitext(os.path.basename(input_file_path))
                for old_suffix in designacao_list:
                    filename = filename.replace("_C." + old_suffix, "")
                output_file_path = os.path.join(
                    folder_path, filename + new_suffix + ".dxf"
                )

                os.rename(input_file_path, output_file_path)

    dxf_files = [
        file for file in os.listdir(folder_path) if file.lower().endswith(".dxf")
    ]

    for dxf_file in dxf_files:
        input_file_path = os.path.join(folder_path, dxf_file)
        process_dxf_file(input_file_path)


def process_dxf_files(input_folder):
    output_folder = os.path.join(input_folder, "Imagens")
    os.makedirs(output_folder, exist_ok=True)

    dxf_files = []
    for root, _, files in os.walk(input_folder):
        dxf_files.extend([os.path.join(root, f) for f in files if f.endswith(".dxf")])

    for dxf_path in dxf_files:
        doc = ezdxf.readfile(dxf_path)
        msp = doc.modelspace()
        base_filename = re.split(
            r"[´]", os.path.splitext(os.path.basename(dxf_path))[0]
        )[0]
        png_path = os.path.join(output_folder, f"{base_filename}.png")

        fig = plt.figure()
        ax = fig.add_axes([0, 0, 1, 1])
        ctx = RenderContext(doc)
        ctx.set_current_layout(msp)
        out = MatplotlibBackend(ax)
        Frontend(ctx, out).draw_layout(msp, finalize=True)

        fig.savefig(png_path, dpi=300)
        plt.close(fig)

    images = []
    filenames = []

    png_files = [f for f in os.listdir(output_folder) if f.endswith(".png")]

    for filename in png_files:
        image_path = os.path.join(output_folder, filename)
        images.append(Image.open(image_path))
        filenames.append(re.split(r"[.`]", filename)[0])

    if images:
        max_width = max(image.width for image in images)
        total_height = sum(image.height for image in images) + len(images) * 40

        combined_image = Image.new("RGB", (max_width, total_height), color="white")
        draw = ImageDraw.Draw(combined_image)
        font = ImageFont.truetype("arial.ttf", 36)

        y_offset = 0
        for image, filename in zip(images, filenames):
            combined_image.paste(image, (0, y_offset))
            text_bbox = draw.textbbox((0, 0), filename, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_position = (
                image.width // 2 - text_width // 2,
                y_offset + image.height,
            )
            draw.text(text_position, filename, font=font, fill=(0, 0, 0))
            y_offset += image.height + 40

        combined_image.save(os.path.join(output_folder, "!DXF_Comp.png"))


# SCRIPT START
process_dxf_folder(folder_path)
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
file_list = os.listdir(folder_path)
data = []
counter = {}
for file_name in file_list:
    if file_name.endswith(".dxf"):
        file_path = os.path.join(folder_path, file_name)
        values = file_name.split("_")
        first_column = "_".join(values[:3]).split("´")[0]

        if first_column in counter:
            counter[first_column] += 1
            modified_first_column = f"{first_column}_{counter[first_column]}"
        else:
            counter[first_column] = 1
            modified_first_column = first_column

        second_column = re.search(r"Mat\.(.*?)_", file_name).group(1)
        third_column = re.search(r"_Thick\.(.*?)_", file_name).group(1)
        fourth_column = re.search(r"_QtA\.(.*?)_", file_name).group(1)

        if "_QtP." in file_name:
            fifth_column = re.search(r"_QtP\.(.*?)\_", file_name).group(1)
        else:
            fifth_column = re.search(r"_QtP\.(.*?)$", file_name).group(1)
        sixth_column = re.search(r"_Dim\.(.*?)_", file_name).group(1)

        seventh_column = re.search(r"_C\.(.*?)$", file_name).group(1)
        seventh_column = seventh_column[:-4]
        if seventh_column in acronimo_mapping:
            seventh_column = acronimo_mapping[seventh_column]

        data.append(
            [
                modified_first_column,
                second_column,
                third_column,
                fourth_column,
                fifth_column,
                sixth_column,
                seventh_column,
            ]
        )

        second_column_folder = os.path.join(output_folder, second_column)
        second_column_folder = second_column_folder.rstrip()
        os.makedirs(second_column_folder, exist_ok=True)
        third_column_folder = os.path.join(second_column_folder, third_column)
        os.makedirs(third_column_folder, exist_ok=True)

        new_file_name = f"{modified_first_column}.dxf"
        new_file_path = os.path.join(third_column_folder, new_file_name)
        shutil.copyfile(file_path, new_file_path)

df = pd.DataFrame(
    data, columns=["Nome", "Mat.", "Esp.", "QtA.", "QtP.", "Dim.", "Corte"]
)
df["Qt."] = df["QtA."].astype(int) * df["QtP."].astype(int)
df = df.drop(["QtA.", "QtP."], axis=1)
df["Qt."] = df["Qt."].astype(int)
df[["Nome", "Mat.", "Esp.", "Dim.", "Corte", "Qt."]] = df[
    ["Mat.", "Esp.", "Nome", "Dim.", "Corte", "Qt."]
]

df.columns = ["Material", "Esp.", "DXF", "Dim.", "Corte", "Quant."]
df["Material"] = df["Material"].str.replace(" ", "")
df["Material"] = df["Material"].str.replace("AlloySteel", "Aisi304")
df["Material"] = df["Material"].str.replace("AISI304", "Aisi304")
df["Material"] = df["Material"].str.replace("1.0037(S235JR)", "S235JR")
df["Material"] = df["Material"].str.replace("PlainCarbonSteel", "S235JR")

light_gray_rgb = "E6E6E6"

start_row = 2
max_content_length = [
    0
] * 5  

print("A gerar nesting...")

df_nl = None
not_in_nesting_list = []

for value in df["Corte"].unique():
    if value not in nesting:
        las_df = df[df["Corte"] == value]
        not_in_nesting_list.append(las_df)

df_nl = pd.concat(not_in_nesting_list)

columns_to_drop = ["DXF", "Corte"]
df_nl = df_nl.drop(columns=columns_to_drop)

mask = df_nl["Quant."] > 1
df_nl = df_nl.loc[df_nl.index.repeat(df_nl["Quant."])]

df_nl.loc[mask, "Quant."] = 1
df_nl = df_nl.sort_values(["Material", "Esp."])
df_nl = df_nl.drop(columns="Quant.")
esp_material_laser = {}
unique_material_laser = df_nl["Material"].unique()

for material in unique_material_laser:
    material_l_df = df_nl[df_nl["Material"] == material].copy()
    esp_l_data = {}
    unique_l_esps = material_l_df["Esp."].unique()
    for esp in unique_l_esps:
        esp_df = material_l_df[material_l_df["Esp."] == esp].copy()
        esp_l_data[esp] = esp_df

    esp_material_laser[material] = esp_l_data
l_dict = {}  
l_list = []


for material, esp_data in esp_material_laser.items():
    for esp, esp_df in esp_data.items():
        dim_values = esp_df[
            "Dim."
        ].tolist()  
        dim_tuples = [tuple(map(int, map(float, dim.split("x")))) for dim in dim_values]

        sorted_dim_tuples = sorted(dim_tuples, key=lambda x: x[0], reverse=True)

        variable_name = f"list_{material}_{esp}" 

        l_dict[variable_name] = sorted_dim_tuples
laser_df_nest = pack_rectangles(l_dict)

laser_df_nest["Codigo"] = ""
for index, row in laser_df_nest.iterrows():
    result = None
    thick = row["Esp."]
    mat = row["Material"]
    result = find_code_sheet(sheet_phc_location, thick, mat)
    if result != None:
        laser_df_nest.at[index, "Codigo"] = result
print(laser_df_nest)

writer = pd.ExcelWriter(output_file, engine="openpyxl")

df.to_excel(writer, sheet_name="DXF", index=False)
unique_mat = df["Material"].unique()

writer.book._sheets.clear()
start_row = 2
max_content_length = [0] * 5

for mat_value in unique_mat:
    filtered_df = df[df["Material"] == mat_value]

    filtered_df = filtered_df.sort_values("Esp.")

    if start_row > 2:
        start_row += 3  
    filtered_df.to_excel(writer, sheet_name="DXF", index=False, startrow=start_row)
    worksheet = writer.sheets["DXF"]
    num_rows = len(filtered_df)

    start_cell = worksheet.cell(row=start_row + 2, column=1)
    end_cell = worksheet.cell(row=start_row + 2 + num_rows - 1, column=6)
    table_range = start_cell.coordinate + ":" + end_cell.coordinate
    table_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in worksheet[table_range]:
        for cell in row:
            cell.border = table_border

    if num_rows > 1:
        start_cell = worksheet.cell(row=start_row + 2, column=1)
        end_cell = worksheet.cell(row=start_row + 2 + num_rows - 1, column=1)
        worksheet.merge_cells(start_cell.coordinate + ":" + end_cell.coordinate)

        start_cell.value = mat_value

        alignment = Alignment(vertical="top")
        font = Font(size=14, bold=True)
        start_cell.alignment = alignment
        start_cell.font = font

    for column in range(1, 7):
        cell = worksheet.cell(row=start_row + 1, column=column)
        cell.fill = PatternFill(fill_type="solid", fgColor=light_gray_rgb)

    if num_rows > 1:
        prev_esp_value = filtered_df["Esp."].iloc[0]
        merge_start_row = start_row + 2
        merge_end_row = start_row + 2
        for row in range(start_row + 2, start_row + 2 + num_rows):
            esp_value = filtered_df["Esp."].iloc[row - start_row - 2]
            if esp_value != prev_esp_value:
                if merge_start_row != merge_end_row:
                    start_cell = worksheet.cell(row=merge_start_row, column=2)
                    end_cell = worksheet.cell(row=merge_end_row, column=2)
                    worksheet.merge_cells(
                        start_cell.coordinate + ":" + end_cell.coordinate
                    )

                    start_cell.value = prev_esp_value
                    alignment = Alignment(vertical="top")
                    font = Font(size=12, bold=True)
                    start_cell.font = font

                merge_start_row = row
                merge_end_row = row
                prev_esp_value = esp_value
            else:
                merge_end_row = row

        if merge_start_row != merge_end_row:
            start_cell = worksheet.cell(row=merge_start_row, column=2)
            end_cell = worksheet.cell(row=merge_end_row, column=2)
            worksheet.merge_cells(start_cell.coordinate + ":" + end_cell.coordinate)
            start_cell.value = prev_esp_value
            alignment = Alignment(vertical="top")
            font = Font(size=12, bold=True)
            start_cell.alignment = alignment
            start_cell.font = font

    for row in range(start_row + 2, start_row + 2 + num_rows):
        for column in range(1, 6):
            cell_value = worksheet.cell(row=row, column=column).value
            if cell_value:
                content_length = len(str(cell_value))
                if content_length > max_content_length[column - 1]:
                    max_content_length[column - 1] = content_length
    start_row += num_rows

for column in range(1, 6):
    col_letter = get_column_letter(column)
    max_length = max_content_length[column - 1]

    min_width = 8

    if column == 1:
        worksheet.column_dimensions[col_letter].width = max(
            min_width, max_length + 5
        )  
    else:
        worksheet.column_dimensions[col_letter].width = max(min_width, max_length + 2)


for column in range(1, 3):
    for row in range(2, start_row + 2):
        cell = worksheet.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", horizontal="center")

worksheet = writer.sheets["DXF"]

worksheet.merge_cells("A1:F1")

merged_cell = worksheet["A1"]
merged_cell.value = ass_name + " - LISTAGEM DE PEÇAS A CORTAR"

border = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
fill = PatternFill(start_color="FFEEE0", end_color="FFEEE0", fill_type="solid")

row_width = 0
for column in range(1, 6):
    row_width += worksheet.column_dimensions[get_column_letter(column)].width
max_font_size = int(row_width * 0.14)  

font_size = max_font_size

font = Font(bold=True, size=font_size)
alignment = Alignment(horizontal="center", vertical="top")
merged_cell.font = font
merged_cell.alignment = alignment
merged_cell.border = border
merged_cell.fill = fill

for row in worksheet["A1:F1"]:
    for cell in row:
        cell.border = border

image_folder = os.path.join(folder_new, "Imagens")

max_row = worksheet.max_row
max_column = worksheet.max_column

column_range = worksheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=max_row)

for column in column_range:
    for cell in column:
        add_image_link(cell, ".png", output_file)

second_sheet = writer.book.create_sheet("DXF_Nest")
first_row = second_sheet.max_row
start_row_laser_df = first_row  
laser_df_nest.to_excel(
    writer, sheet_name="DXF_Nest", index=False, startrow=start_row_laser_df
)
value_cell_nest = f"Nesting - {ass_name}"
nesting_title_cell = second_sheet.cell(row=first_row, column=1, value=value_cell_nest)
nesting_title_cell.alignment = Alignment(horizontal="center", vertical="center")

second_sheet.merge_cells(
    start_row=first_row, start_column=1, end_row=first_row, end_column=4
)
for row in second_sheet.iter_rows(
    min_row=first_row, max_row=first_row, min_col=1, max_col=4
):
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
for row in second_sheet.iter_rows(
    min_row=1, min_col=1, max_row=second_sheet.max_row, max_col=second_sheet.max_column
):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

max_lengths = laser_df_nest.iloc[:, 1:].map(lambda x: len(str(x))).max()

for i, max_length in enumerate(max_lengths):
    column_letter = get_column_letter(i + 2)  
    second_sheet.column_dimensions[column_letter].width = max_length + 8

sum_abcd_width = 0
for column_letter in ["A", "B", "C"]:
    sum_abcd_width += second_sheet.column_dimensions[column_letter].width

if sum_abcd_width < (len(ass_name) + 12):
    additional_width = (len(ass_name) + 12) - sum_abcd_width
    equal_width = additional_width / 3  
    for column_letter in ["A", "B", "C"]:
        second_sheet.column_dimensions[column_letter].width += equal_width
writer.book.save(output_file)

process_dxf_files(folder_new)
convert_worksheet_to_pdf(output_file, worksheet_name, output_pdf)
put_watermark(output_pdf, output_pdf, logo_image_path)
add_page_numbers(output_pdf)
try:
    convert_worksheet_to_pdf(output_file, worksheet_name_l, output_nesting_laser_pdf)
    put_watermark(output_nesting_laser_pdf, output_nesting_laser_pdf, logo_image_path)
    add_page_numbers(output_nesting_laser_pdf)
except:
    pass

if edit_val == "0":
    showMessage("Processo Terminado")

sys.exit()
