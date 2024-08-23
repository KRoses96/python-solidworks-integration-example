"""
Author: Manuel Matias Rosa
Description: Creates a list and organizes/analyses the data for all the weldment profiles used in the assembly
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
import tkinter as tk
from tkinter import messagebox
import sys
import math
import openpyxl
from PyPDF2 import PdfMerger
from PyPDF4 import PdfFileWriter, PdfFileReader
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from win32com import client
import io
import shutil
import tkinter as tk
from tkinter import Scrollbar, Text, Frame
import math
from unidecode import unidecode
from profile_phc_checker import code_finder


# FUNCTIONS
def create_excel(df, title, output_filename):
    if getattr(sys, "frozen", False):
        base_path = sys._MEIPASS
    else:
        base_path = ""

    output_path = os.path.join(base_path, output_filename)

    writer = pd.ExcelWriter(output_path)
    df.to_excel(writer, sheet_name="Perfis", index=False, startrow=1)

    workbook = writer.book
    worksheet = writer.sheets["Perfis"]

    title_format = workbook.add_format(
        {
            "bold": True,
            "text_wrap": True,
            "valign": "top",
            "align": "center",
            "fg_color": "#FFD700",  
            "font_color": "#000000",  
            "font_size": 14,
            "border": 1, 
        }
    )
    worksheet.write(0, 0, title, title_format)
    worksheet.merge_range(
        "A1:F1", title, title_format
    )  

    header_format = workbook.add_format(
        {
            "bold": True,
            "text_wrap": True,
            "valign": "top",
            "align": "center",
            "fg_color": "#121263",
            "font_color": "#FFFFFF",
            "border": 1,  
        }
    )

    for col_num, value in enumerate(df.columns.values):
        worksheet.write(
            1, col_num, value, header_format
        )
    center_alignment = workbook.add_format(
        {"align": "center", "valign": "vcenter", "border": 1}
    )
    left_aligned_format = workbook.add_format(
        {"align": "left", "valign": "vcenter", "border": 1}
    )

    for col_num, col_name in enumerate(df.columns):
        col_letter = chr(ord("A") + col_num)
        max_len = max(df[col_name].astype(str).apply(len).max(), len(col_name))
        if col_num == 0:
            worksheet.set_column(
                col_letter + ":" + col_letter,
                max_len + 2,
                cell_format=left_aligned_format,
            )
        else:
            worksheet.set_column(
                col_letter + ":" + col_letter, max_len + 2, cell_format=center_alignment
            )
    writer.close()


def showMessage(message, type="info", timeout=5000):
    import tkinter as tk
    from tkinter import messagebox as msgb

    root = tk.Tk()
    root.withdraw()
    try:
        root.after(timeout, root.destroy)
        if type == "info":
            msgb.showinfo("Terminado", message, master=root)
        elif type == "warning":
            msgb.showwarning("Warning", message, master=root)
        elif type == "error":
            msgb.showerror("Error", message, master=root)
    except:
        pass


def merge_pdfs(folder_path, output_path):
    merger = PdfMerger()

    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            filepath = os.path.join(folder_path, filename)
            with open(filepath, "rb") as file:
                merger.append(file)
    with open(output_path, "wb") as output_file:
        merger.write(output_file)
    shutil.rmtree(folder_path)
    merger.close()



def put_watermark(input_pdf, output_pdf, watermark):
    watermark_instance = PdfFileReader(watermark)
    watermark_page = watermark_instance.getPage(0)


    scale_factor = 1 / 8  
    scaled_watermark_height = watermark_page.mediaBox.getUpperRight_y() * scale_factor
    watermark_page.scaleBy(scale_factor)
    pdf_reader = PdfFileReader(input_pdf)
    pdf_writer = PdfFileWriter()
    
    for page_num in range(pdf_reader.getNumPages()):
        page = pdf_reader.getPage(page_num)

        page_width = page.mediaBox.getUpperRight_x()
        page_height = page.mediaBox.getUpperRight_y()
        page_width = float(page_width)
        page_height = float(page_height)

        x = 47  
        y = (
            page_height - scaled_watermark_height
        )  
        page.mergeTranslatedPage(watermark_page, x, y)

        pdf_writer.addPage(page)

    with open(output_pdf, "wb") as out:
        pdf_writer.write(out)


def create_page_pdf_2(num):
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)

    for i in range(1, num + 1):
        text = f"{i}/{num}"

        text_width = c.stringWidth(text, "Helvetica", 12)
        x = (210 - text_width) * mm  
        y = 4 * mm  

        c.drawString(x, y, text)
        c.showPage()

    c.save()

    packet.seek(0)
    return PdfFileReader(packet)


def create_page_pdf(page_number, total_pages):
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)

    for i in range(1, total_pages + 1):
        text = f"{page_number}/{total_pages}"
        text_width = c.stringWidth(text, "Helvetica", 12)
        x = (300 - text_width) * mm  
        y = 4 * mm  
        c.drawString(x, y, text)
        c.showPage()

    c.save()

    packet.seek(0)
    return PdfFileReader(packet)


def add_page_numbers_2(pdf_path):
    """
    Add page numbers to a pdf, save the result as a new pdf
    @param pdf_path: path to pdf
    """
    output_pdf = PdfFileWriter()
    with open(pdf_path, "rb") as f:
        reader = PdfFileReader(f, strict=False)
        n = reader.getNumPages()

        number_pdf = create_page_pdf_2(n)

        for p in range(n):
            page = reader.getPage(p)
            number_layer_page = number_pdf.getPage(p)

            page.mergePage(number_layer_page)
            output_pdf.addPage(page)
        if output_pdf.getNumPages():
            newpath = pdf_path[:-4] + "_numbered.pdf"
            with open(newpath, "wb") as f:
                output_pdf.write(f)
        shutil.move(newpath, pdf_path)


def add_page_numbers(pdf_path, page_number, total_pages):
    """
    Add page numbers to a pdf, save the result as a new pdf
    @param pdf_path: path to pdf
    """
    output_pdf = PdfFileWriter()

    with open(pdf_path, "rb") as f:
        reader = PdfFileReader(f, strict=False)
        n = reader.getNumPages()

        number_pdf = create_page_pdf(page_number, total_pages)

        for p in range(n):
            page = reader.getPage(p)
            number_layer_page = number_pdf.getPage(p)
            page.mergePage(number_layer_page)
            output_pdf.addPage(page)

        if output_pdf.getNumPages():
            newpath = os.path.splitext(pdf_path)[0] + f"_numbered.pdf"
            with open(newpath, "wb") as f:
                output_pdf.write(f)
        shutil.move(newpath, pdf_path)


def add_hyperlink(sheet, cell, relative_file_path, original_text):
    sheet[cell].hyperlink = f"{relative_file_path}"
    sheet[cell].value = original_text
    sheet[cell].font = Font(color="0000FF", underline="single")


def search_for_file(directory_path, base_name, extensions):
    for root, dirs, files in os.walk(directory_path):
        for ext in extensions:
            file_path = os.path.join(root, f"{base_name}.{ext}")
            if os.path.exists(file_path):
                return file_path
    return None


def opt_cut(df, max_value):
    df = df.reindex(df.index.repeat(df["QT"]))
    df = df.drop(columns=["QT"])
    df = df.sort_values(by="LEN", ascending=False)
    error_message = ""
    remaining_values = [max_value]
    for index, row in df.iterrows():
        len_value = row["LEN"]
        if len_value > max_value:
            error_message = "ERRO: Valor > 6metros"
            continue

        closest_value = None
        min_difference = float("inf")
        for i, value in enumerate(remaining_values):
            difference = abs(value - len_value)
            if difference < min_difference and value >= len_value:
                min_difference = difference
                closest_value = value
                closest_index = i
        if max(remaining_values) < len_value:
            closest_value = max_value
            last_index = len(remaining_values)
            closest_index = last_index
            remaining_values.append(max_value)
        remaining_values.pop(closest_index)
        remaining_values.append(closest_value - len_value)
    Opt_sum = len(remaining_values) * max_value
    return Opt_sum, error_message


def show_error_row(row_data, error_message):
    root = tk.Tk()
    root.title("Erro Perfil")
    root.geometry("500x400")

    frame = Frame(root)
    frame.pack(fill=tk.BOTH, expand=1)
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    xscrollbar = Scrollbar(frame, orient=tk.HORIZONTAL)
    xscrollbar.grid(row=1, column=0, sticky=tk.E + tk.W)
    yscrollbar = Scrollbar(frame)
    yscrollbar.grid(row=0, column=1, sticky=tk.N + tk.S)

    text = Text(
        frame,
        wrap=tk.NONE,
        xscrollcommand=xscrollbar.set,
        yscrollcommand=yscrollbar.set,
    )
    text.grid(row=0, column=0, sticky="nsew")
    xscrollbar.config(command=text.xview)
    yscrollbar.config(command=text.yview)

    text.insert(tk.END, f"Erro no seguinte item:\n")
    for key, value in row_data.items():
        text.insert(tk.END, f"{key}: {value}\n")
    text.insert(tk.END, f"Erro: {error_message}")
    root.mainloop()



def remove_trailing_zeroes(input_string):
    pattern = r"(\d+\.\d*?)0*(?![\d\.])|(\d+\.)"
    result = re.sub(pattern, r"\1\2", input_string)
    return result


def remove_empty_decimal_points(input_string):
    pattern = r"(?<=\d)\.(?![\d])"
    result = re.sub(pattern, "", input_string)
    return result


def excel_to_dfs(input_file, output_file):
    df = pd.read_excel(input_file, engine="openpyxl")
    df = df.map(lambda x: str(x).replace("\n", ""))
    df = df.map(lambda x: str(x).replace(",", "."))
    titulo_dict = {}
    previous_titulo = ""
    try:

        def update_item_no(item_no):
            if isinstance(item_no, str) and item_no.endswith(".0"):
                return item_no.rstrip(".0")
            return item_no 

        df["ITEM NO."] = df["ITEM NO."].apply(update_item_no)
        for index, row in df.iterrows():
            item_number = row["ITEM NO."]
            titulo = row["Título"]

            if str(item_number).startswith(
                ("0", "1", "2", "3", "4", "5", "6", "7", "8", "9")
            ):
                if item_number not in titulo_dict:
                    titulo_dict[item_number] = titulo
                    previous_titulo = titulo 

                else:
                    titulo_dict[item_number] = titulo
                    previous_titulo = titulo  

            if (
                "." in str(item_number)
                and item_number[: item_number.rindex(".")] in titulo_dict
            ):
                titulo_dict[item_number] = titulo_dict[
                    item_number[: item_number.rindex(".")]
                ]
                previous_titulo = titulo_dict[item_number]  

            if row["Título"] == "nan":
                df.loc[index, "Título"] = previous_titulo
        df["QTY."] = df["QTY."].astype(float)
        df["QTY."] = df["QTY."].astype(int)

        df["QTY."] = pd.to_numeric(df["QTY."], errors="coerce").fillna(1)
        df["Multiplied QTY."] = df["QTY."]
        df["ITEM NO."] = df["ITEM NO."].astype("str")

        def calculate_multiplied_qty(item_no, qty):
            parent_item_no = ".".join(item_no.split(".")[:-1])
            if parent_item_no:
                parent_row = df[df["ITEM NO."] == parent_item_no]
                if not parent_row.empty:
                    parent_qty = parent_row.iloc[0]["Multiplied QTY."]
                    return qty * parent_qty
            return qty

        for index, row in df.iterrows():
            df.at[index, "Multiplied QTY."] = calculate_multiplied_qty(
                row["ITEM NO."], row["QTY."]
            )
        pd.set_option("display.max_colwidth", None)
        pd.set_option("display.max_rows", None)
        df = df.sort_index()

        df = df.drop("QTY.", axis=1)
        df = df.dropna(subset=["ANGLE1", "ANGLE2", "LENGTH"], how="any")
        df = df[
            (df["ANGLE1"] != "nan") & (df["ANGLE2"] != "nan") & (df["LENGTH"] != "nan")
        ]
        df = df[df["ANGLE1"] != ""]
        df = df[df["ANGLE2"] != ""]
        df = df[df["LENGTH"] != ""]
        df = df[df["LENGTH"] != "nan"]
        df["ANGLE1"] = df["ANGLE1"].apply(lambda x: re.sub("[^0-9.\-]", "", str(x)))
        df["ANGLE2"] = df["ANGLE2"].apply(lambda x: re.sub("[^0-9.\-]", "", str(x)))
        pd.set_option("display.max_columns", None)
        df.loc[df["Description"] == "nan", "Description"] = df["Designação"]
        df.loc[df["Description"] == "", "Description"] = df["Designação"]

        df["Description"] = df["Description"].str.replace("X", "x")
        df["Description"] = df["Description"].str.replace("x ", "x")
        df["Description"] = df["Description"].str.replace(" x", "x")
        df["Description"] = df["Description"].str.replace(" x ", "x")
        df["Description"] = df["Description"].apply(remove_trailing_zeroes)
        df["Description"] = df["Description"].apply(remove_empty_decimal_points)
        df["MATERIAL"] = df["MATERIAL"].str.replace(" ", "")

        df["MATERIAL"] = df["MATERIAL"].str.replace("1.0037(S235JR)", "S235JR")
        df["MATERIAL"] = df["MATERIAL"].str.replace("PlainCarbonSteel", "S235JR")
        df["MATERIAL"] = df["MATERIAL"].str.replace("AISI304", "Aisi304")
        df["MATERIAL"] = df["MATERIAL"].str.replace("AlloySteel", "Aisi304")
        if len(df) == 0:
            sys.exit()
        df = df[df["Chapa"] == "nan"]
        df = df.iloc[:, 3:]
        df.columns = ["DES", "MAT", "ANG1", "ANG2", "LEN", "TIT", "QT"]

        df["ANG1"] = df["ANG1"].replace("-", "0.0")
        df["ANG2"] = df["ANG2"].replace("-", "0.0")
        df["ANG1"] = df["ANG1"].replace("-", "0.0")
        df["ANG2"] = df["ANG2"].replace("-", "0.0")
        df["ANG1"] = df["ANG1"].astype(float)
        df["ANG2"] = df["ANG2"].astype(float)

        df["QT"] = df["QT"].astype(float)
        df["QT"] = df["QT"].astype(int)
        df["LEN"] = df["LEN"].astype(float)
        df["TIT"] = df["TIT"].replace("nan", "")
        pd.set_option("display.max_columns", None)
        pd.set_option("display.max_rows", None)
        df["LEN"] = df["LEN"].round(0)
        df = (
            df.groupby(["DES", "MAT", "LEN", "TIT"], as_index=False)
            .apply(
                lambda x: (
                    x
                    if ((x["ANG1"] > 0) | (x["ANG2"] > 0)).any()
                    else x.groupby(
                        ["DES", "MAT", "ANG1", "ANG2", "LEN", "TIT"], as_index=False
                    )["QT"].sum()
                )
            )
            .reset_index(drop=True)
        )
        dfs = {}
        for DES, MAT in df[["DES", "MAT"]].drop_duplicates().itertuples(index=False):
            dfs[(DES, MAT)] = df[(df["DES"] == DES) & (df["MAT"] == MAT)].reset_index(
                drop=True
            )
    except Exception as e:
        show_error_row(row, str(e))
        raise  

    workbook = Workbook()

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    counter_DES = 0
    df_final = df
    for (DES, MAT), df in dfs.items():
        counter_DES += 1
        sheet_name = f"{counter_DES}"
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.insert_rows(0, 4)
        sheet["A1"] = "LISTAGEM DE PERFIS"
        sheet["A2"] = " "
        sheet["A3"] = " "

        sheet["A2"] = file_name_1

        sheet["A4"] = " "

        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

        gray_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
        orange_fill = PatternFill(
            start_color="FFEEE0", end_color="FFEEE0", fill_type="solid"
        )

        sheet["A2"].fill = orange_fill
        sheet["B2"].fill = orange_fill
        sheet["A1"].fill = orange_fill

        sheet["A5"] = "Designação"
        sheet["B5"] = "Material"
        sheet["C5"] = "Âng.1"
        sheet["D5"] = "Âng.2"
        sheet["E5"] = "Comprimento (mm)"
        sheet["F5"] = "Conjunto"
        sheet["G5"] = "Quant."
        sheet["H5"] = "  Observações  "

        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment("center")
        for cell in sheet[2]:
            cell.font = Font(bold=True)
        for cell in sheet[3]:
            cell.font = Font(bold=True)
        for cell in sheet[5]:
            cell.font = Font(bold=True)
            cell.fill = gray_fill

        for row in sheet.iter_rows(
            min_row=5, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column
        ):
            for cell in row:
                cell.border = thin_border

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for i, cell in enumerate(column):
                if i < 3:  
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        sum_abcd_width = 0
        for column_letter in ["A", "B", "C", "D"]:
            sum_abcd_width += sheet.column_dimensions[column_letter].width

        if sum_abcd_width < (len(file_name_1) + 2):
            additional_width = (len(file_name_1) + 2) - sum_abcd_width

            equal_width = additional_width / 4  
            for column_letter in ["A", "B", "C", "D"]:
                sheet.column_dimensions[column_letter].width += equal_width

        sum_row = [""]
        sheet.append(sum_row)
        workbook_remove = None
        units_sum = None
        sum_row = ["Comprimento Total:", None, None, None, None, None]
        sheet.append(sum_row)
        matching_keyword = None  
        for keyword_m in keys_list:
            if keyword_m in unidecode(DES.lower()):
                matching_keyword = keyword_m
                condition = filter_conditions[keyword_m]
                if (
                    condition["remove"] == 0
                    and math.isnan(condition["profile"]) == True
                    or optimi == "0"
                ):
                    df["TOTAL_LEN"] = df["LEN"] * df["QT"] * 0.001
                    sum_total = round(df["TOTAL_LEN"].sum(), 4)
                elif condition["remove"] == 0 and optimi == "1":
                    error_message = opt_cut(df, condition["profile"])[1]
                    result_sum = opt_cut(df, condition["profile"])[0] * 0.001
                    if int(result_sum / (condition["profile"] * 0.001)) == 1:
                        units_sum = (
                            "|"
                            + str(int(result_sum / (condition["profile"] * 0.001)))
                            + " Unidade "
                        )
                    else:
                        units_sum = (
                            " | "
                            + str(int(result_sum / (condition["profile"] * 0.001)))
                            + " Unidades "
                        )
                    if error_message == "":
                        sum_total = result_sum
                    else:
                        max_opti = condition["profile"]
                        sum_total = f"ERRO: Valores > {max_opti}"

                else:
                    workbook_remove = 1
                    sum_total = "FOLHA PARA SER APAGADA"

                break  

        if matching_keyword is None:
            df["TOTAL_LEN"] = df["LEN"] * df["QT"] * 0.001
            sum_total = round(df["TOTAL_LEN"].sum(), 4)
        sheet.merge_cells(
            start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=2
        )

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells(
            start_row=sheet.max_row, start_column=3, end_row=sheet.max_row, end_column=5
        )
        merged_cell = sheet.cell(row=sheet.max_row, column=3)
        if units_sum == None:
            merged_cell.value = str(sum_total) + " m"
        else:
            merged_cell.value = str(sum_total) + " m" + units_sum
        comprimento_total_cell = sheet.cell(row=sheet.max_row, column=1)
        comprimento_total_cell.alignment = Alignment(horizontal="center")

        total_len_cell = sheet.cell(row=sheet.max_row, column=3)
        total_len_cell.alignment = Alignment(horizontal="left")

        blue_fill = PatternFill(
            start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
        )
        comprimento_total_cell.fill = blue_fill
        total_len_cell.fill = blue_fill

        bold_font = Font(bold=True, size=14)
        comprimento_total_cell.font = bold_font
        total_len_cell.font = bold_font

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in sheet.iter_rows(
            min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=5
        ):
            for cell in row:
                cell.border = border
        a2 = sheet["A2"]
        c2 = sheet["C2"]

        sheet.merge_cells("A1:D1")
        sheet.merge_cells("A2:D2")
        darker_blue_fill = PatternFill(
            start_color="A9CCE3", end_color="A9CCE3", fill_type="solid"
        )
        for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
            for cell in row:
                cell.fill = darker_blue_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        for cell in [a2, c2]:
            cell.fill = orange_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in range(5, sheet.max_row - 1):
            sheet.row_dimensions[row].height = 30
        if workbook_remove == 1:
            workbook.remove(sheet)
        code = None
        if code != None:
            code_row = [""]
            sheet.append(code_row)
            code_row = [f"Código : {code}"]
            sheet.append(code_row)
            for cell in sheet[sheet.max_row]:
                cell.font = Font(bold=True, underline="single", size=12)

        for row_num in range(6, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=6).value
            if cell_value:
                file_extensions = ["SLDPRT", "SLDASM"]
                file_path = search_for_file(
                    parent_directory, cell_value, file_extensions
                )
                if file_path:
                    relative_path = os.path.relpath(
                        file_path, os.path.dirname(output_file)
                    )
                    add_hyperlink(sheet, f"F{row_num}", relative_path, cell_value)

    if "Sheet" in workbook.sheetnames:
        sheet = workbook["Sheet"]

        workbook.remove(sheet)
    workbook.save(output_file)
    return df_final


# SCRIPT START
script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
parent_dir = os.path.dirname(script_dir)
options = os.path.join(parent_dir, "extras", "op.txt")
with open(options, "r") as file:
    lines = file.readlines()
if lines:
    phc_data = lines[9].strip()
    if phc_data:
        phc_path = (((phc_data).split("="))[1])[1:]
print(phc_path)

csv_file = os.path.join(parent_dir, "extras", "cutlist.csv")
excel_file = os.path.join(phc_path, "aproveitamento.xlsx")
df_filter = pd.read_excel(excel_file)
filter_conditions = {}
for index, row in df_filter.iterrows():
    keywords = [keyword.strip() for keyword in row["Filtro"].split(";")]
    profile = row["Comp_Perfil"]
    remove = row["Remover"]
    for keyword in keywords:
        filter_conditions[keyword] = {"profile": profile, "remove": remove}
keys_list = list(filter_conditions.keys())
file_name = os.path.join(parent_dir, "macros", "file_path.txt")
source_name = os.path.join(parent_dir, "macros", "source_path.txt")

with open(file_name, "r") as file:
    file_path = file.read().strip()

with open(source_name, "r") as file:
    source_path = file.read().strip()

file_name = os.path.basename(file_path)

data = pd.read_csv(file_path, delimiter="\t", encoding="latin-1")

file_path_txt = file_path
file_path = file_path[:-3] + "xlsx"
file_name = os.path.basename(file_path)
parent_directory = os.path.dirname(source_path)
writer = pd.ExcelWriter(file_path, engine="xlsxwriter")

data.to_excel(writer, index=False, sheet_name="Perfis")

writer.close()
file_name_1 = file_name[0:-5]

input_file = file_path 

parent_directory = os.path.dirname(source_path)
filename = os.path.basename(source_path)
filename_without_extension = os.path.splitext(filename)[0]

options = os.path.join(parent_dir, "extras", "op.txt")

with open(options, "r") as file:
    lines = file.readlines()

if lines:
    first_line = lines[
        0
    ].strip()  
    if first_line:
        macro_to_run = first_line[-1]

if lines:
    fourth_line = lines[
        3
    ].strip()  
    if first_line:
        optimi = fourth_line[-1]  

if lines:
    phc_data_line = lines[9].strip()
    if phc_data_line:
        phc_data_path = (((phc_data_line).split("="))[1])[1:]
phc_data_path = os.path.join(phc_data_path, "Perfis")
new_folder_name = "Listas_" + filename_without_extension

new_folder_path = os.path.join(parent_directory, new_folder_name)
if not os.path.exists(new_folder_path):
    os.makedirs(new_folder_path)

default_file_name = f"\\ {filename_without_extension}" + "_Perfis_Corte.xlsx"
default_file_name_res = f"\\ {filename_without_extension}" + "_Perfis.xlsx"
file_name_final = new_folder_path + default_file_name
output_file = file_name_final
output_file_res = new_folder_path + default_file_name_res
export_path = os.path.join(new_folder_path, f"{filename_without_extension}_Perfis.pdf")
if os.path.exists(output_file_res):
    try:
        os.remove(output_file_res)
        os.remove(export_path)
    except:
        pass

df_final = excel_to_dfs(input_file, output_file)
df_final = df_final.drop("ANG1", axis=1)
df_final = df_final.drop("ANG2", axis=1)
df_final = df_final.drop("TIT", axis=1)
if len(df_final) == 0:
    sys.exit()
df_final["LEN_QTY"] = df_final["LEN"] * df_final["QT"]
df_final["LEN_QTY"] = pd.to_numeric(df_final["LEN_QTY"], errors="coerce")
df_final["Comp. Compra"] = ""
df_final["Unidades"] = ""
matching_keyword = None
for group, group_df in df_final.groupby(["DES", "MAT"]):
    for keyword_m in keys_list:
        if keyword_m in unidecode(group[0].lower()):
            matching_keyword = keyword_m
            condition = filter_conditions[keyword_m]
            if (
                condition["remove"] == 0
                and math.isnan(condition["profile"]) == True
                or optimi == "0"
            ):
                pass
            elif condition["remove"] == 0 and optimi == "1":
                result_sum = opt_cut((group_df), condition["profile"])[0] * 0.001
                units_sum = str(int(result_sum / (condition["profile"] * 0.001)))
                result_sum = str(result_sum) + f"({str(condition['profile']*0.001)}m)"
                df_final.loc[group_df.index, "Unidades"] = units_sum
                df_final.loc[group_df.index, "Comp. Compra"] = result_sum
            break  
    if matching_keyword is None:
        pass

df_final = df_final.drop("LEN", axis=1)
df_final = df_final.drop("QT", axis=1)
df_final = df_final.groupby(["DES", "MAT", "Comp. Compra", "Unidades"], as_index=False)[
    "LEN_QTY"
].sum()
df_final = df_final.rename(
    columns={
        "DES": "Descrição",
        "MAT": "Material",
        "LEN_QTY": "Comp.(m)",
        "Comp. Compra": "Comp.C.(m)",
        "Unidades": "Qt.",
    }
)
df_final["Comp.(m)"] = pd.to_numeric(df_final["Comp.(m)"], errors="coerce").mul(0.001)
df_final = df_final[["Descrição", "Material", "Comp.(m)", "Comp.C.(m)", "Qt."]]
df_final["Codigo"] = ""
df_final["Comp.(m)"] = df_final["Comp.(m)"].round(4)

for index, row in df_final.iterrows():
    descri = row["Descrição"]
    material = row["Material"]
    result = code_finder(descri, material, phc_data_path)
    df_final.at[index, "Codigo"] = result
title = f"{filename_without_extension}" + " Perfis"
PDF_folder_path = os.path.dirname(output_file) + "\\Pdf"
workbook = openpyxl.load_workbook(output_file)
workbook.save(output_file)
if not os.path.exists(PDF_folder_path):
    os.makedirs(PDF_folder_path)
excel = client.Dispatch("Excel.Application")

try:
    workbook = excel.Workbooks.Open(file_name_final)
except:
    window = tk.Tk()
    window.withdraw()
    messagebox.showerror(
        "Erro",
        "PDFs não foram criados!\nExcel não foi revisto ou guardado após a revisão!",
    )
    sys.exit()

folder_path = PDF_folder_path
os.makedirs(folder_path, exist_ok=True)

rows_per_page = 25
counter = 0

for worksheet in workbook.Worksheets:
    columns_range = worksheet.Range("A:J")

    xlPaperA4 = 9  
    xlLandscape = 2  
    worksheet.PageSetup.PaperSize = xlPaperA4
    worksheet.PageSetup.Orientation = xlLandscape

    total_rows = worksheet.UsedRange.Rows.Count
    total_columns = worksheet.UsedRange.Columns.Count

    rows_per_page = 25
    total_pages = (total_rows) / rows_per_page
    if total_pages == int(total_pages):
        rows_per_page = 25
        total_pages = int(total_pages)

    else:
        rows_per_page = round((total_rows / round(total_pages + 0.5)) + 0.5)
        total_pages = round(total_pages + 0.5)

    for page_number in range(1, total_pages + 1):
        counter = counter + 1
        start_row = (page_number - 1) * rows_per_page + 1
        end_row = min((page_number) * rows_per_page, total_rows)

        print_area = worksheet.Range(
            worksheet.Cells(start_row, 1), worksheet.Cells(end_row, total_columns)
        )

        worksheet.PageSetup.Zoom = False
        worksheet.PageSetup.FitToPagesWide = 1

        worksheet.PageSetup.PrintArea = print_area.Address
        pdf_export_format = 0 
        export_path = os.path.join(folder_path, f"{counter}_Page{page_number}.pdf")
        worksheet.ExportAsFixedFormat(pdf_export_format, export_path)
        add_page_numbers(export_path, page_number, total_pages)

workbook.Close(SaveChanges=False)

excel.Quit()
window = tk.Tk()
window.withdraw()
modified_filename_PDF = file_name_final[:-5] + ".pdf"
logo_image_path = os.path.join(parent_dir, "extras", "Logo.pdf")
merge_pdfs(folder_path, modified_filename_PDF)
put_watermark(modified_filename_PDF, modified_filename_PDF, logo_image_path)

create_excel(df_final, title, output_file_res)
try:
    workbook = excel.Workbooks.Open(output_file_res)
except:
    pass
for worksheet in workbook.Worksheets:
    used_range = worksheet.UsedRange
    num_rows = used_range.Rows.Count
    num_cols = used_range.Columns.Count
    start_row = 1
    export_range = worksheet.Range(
        worksheet.Cells(start_row, 1), worksheet.Cells(num_rows, num_cols)
    )
    columns_range = export_range
    worksheet.PageSetup.PrintArea = columns_range.Address
    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = False

    pdf_export_format = 0  
    export_path = os.path.join(
        new_folder_path, f"{filename_without_extension}_Perfis.pdf"
    )
    export_range.ExportAsFixedFormat(pdf_export_format, export_path)

workbook.Save()
excel.Quit()
put_watermark(export_path, export_path, logo_image_path)
add_page_numbers_2(export_path)

if macro_to_run == "1":
    showMessage("Processo Terminado")

sys.exit()
