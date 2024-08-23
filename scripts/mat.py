"""
Author: Manuel Matias Rosa
Description: Creates a list of every item from an assembly that has the property "Paraf"
"""

import pandas as pd
from openpyxl import Workbook, utils
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import messagebox, Frame, Scrollbar, Text
import sys
import win32com.client
import re
from PyPDF4 import PdfFileWriter, PdfFileReader
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import io
import shutil
import numpy as np


pd.set_option("display.max_columns", None)
script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
parent_dir = os.path.dirname(script_dir)
file_name = os.path.join(parent_dir, "macros", "file_path_mat.txt")
source_name = os.path.join(parent_dir, "macros", "source_path_mat.txt")
options = os.path.join(parent_dir, "extras", "op.txt")


def showMessage(message, type="info", timeout=5000):
    import tkinter as tk
    from tkinter import messagebox as msgb

    root = tk.Tk()
    root.withdraw()
    try:
        root.after(timeout, root.destroy)
        if type == "info":
            msgb.showinfo("Terminado", message, master=root)
        elif type == "warning":
            msgb.showwarning("Warning", message, master=root)
        elif type == "error":
            msgb.showerror("Error", message, master=root)
    except:
        pass


def put_watermark(input_pdf, output_pdf, watermark):
    watermark_instance = PdfFileReader(watermark)
    watermark_page = watermark_instance.getPage(0)

    scale_factor = 1 / 8  
    scaled_watermark_width = watermark_page.mediaBox.getUpperRight_x() * scale_factor
    scaled_watermark_height = watermark_page.mediaBox.getUpperRight_y() * scale_factor


    watermark_page.scaleBy(scale_factor)
    pdf_reader = PdfFileReader(input_pdf)
    pdf_writer = PdfFileWriter()
    for page_num in range(pdf_reader.getNumPages()):
        page = pdf_reader.getPage(page_num)
        page_width = page.mediaBox.getUpperRight_x()
        page_height = page.mediaBox.getUpperRight_y()

        page_width = float(page_width)
        page_height = float(page_height)
        x = 47  
        y = (
            page_height - scaled_watermark_height
        )  
        page.mergeTranslatedPage(watermark_page, x, y)
        pdf_writer.addPage(page)
    with open(output_pdf, "wb") as out:
        pdf_writer.write(out)

def create_page_pdf(num):
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)

    for i in range(1, num + 1):
        text = f"{i}/{num}"

        text_width = c.stringWidth(text, "Helvetica", 12)
        x = (210 - text_width) * mm 
        y = 4 * mm  

        c.drawString(x, y, text)
        c.showPage()

    c.save()

    packet.seek(0)
    return PdfFileReader(packet)


def add_page_numbers(pdf_path):
    """
    Add page numbers to a pdf, save the result as a new pdf
    @param pdf_path: path to pdf
    """
    output_pdf = PdfFileWriter()

    with open(pdf_path, "rb") as f:
        reader = PdfFileReader(f, strict=False)
        n = reader.getNumPages()

        number_pdf = create_page_pdf(n)

        for p in range(n):
            page = reader.getPage(p)
            number_layer_page = number_pdf.getPage(p)

            page.mergePage(number_layer_page)
            output_pdf.addPage(page)

        if output_pdf.getNumPages():
            newpath = pdf_path[:-4] + "_numbered.pdf"
            with open(newpath, "wb") as f:
                output_pdf.write(f)
        shutil.move(newpath, pdf_path)

def show_error_row(row_data, error_message):
    root = tk.Tk()
    root.title("Erro Lista Material")


    frame = Frame(root)
    frame.pack(fill=tk.BOTH, expand=1)
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    xscrollbar = Scrollbar(frame, orient=tk.HORIZONTAL)
    xscrollbar.grid(row=1, column=0, sticky=tk.E + tk.W)
    yscrollbar = Scrollbar(frame)
    yscrollbar.grid(row=0, column=1, sticky=tk.N + tk.S)

    text = Text(
        frame,
        wrap=tk.NONE,
        xscrollcommand=xscrollbar.set,
        yscrollcommand=yscrollbar.set,
    )
    text.grid(row=0, column=0, sticky="nsew")
    xscrollbar.config(command=text.xview)
    yscrollbar.config(command=text.yview)

    text.insert(tk.END, f"Erro: {error_message}")

    root.mainloop()


def excel_to_dfs(input_file, output_file):
    df = pd.read_excel(input_file, engine="openpyxl")

    df = df.map(lambda x: str(x).replace("\n", ""))
    df = df.map(lambda x: str(x).replace(",", "."))

    try:
        df.loc[df["Paraf"] == "nan", "Paraf"] = df["ParafC"]
        df.loc[df["Paraf"] == "", "Paraf"] = df["ParafC"]
        mask = (df["ParafC"] != "nan") & (df["ParafC"] != "")
        df.loc[mask, "Paraf"] = df.loc[mask, "ParafC"]
        df["Path"] = df["SW-Folder Name(Folder Name)"] + df["SW-File Name(File Name)"]
        

        df_impressao = pd.DataFrame()
        for index, row in df.iterrows():
            if row['Material'] == 'Impressao':
                df_impressao = df_impressao._append(row)
        print(df_impressao)
        if len(df_impressao) != 0:
            selected_columns = ["ITEM NO.","Qty","Mass","Material"]
            df_impressao = df_impressao[selected_columns]
            df_impressao.replace("nan", np.nan, inplace=True)
            df_impressao = df_impressao.dropna()
            printed_mass = 0.0
            for index, row in df_impressao.iterrows():
                printed_mass = printed_mass + (float(row['Mass']) * float(row['Qty']))
            print_row = pd.Series([f"Filamento de Impressão 3D - {printed_mass}Kg",1,""] , index=["Paraf", "Qty", "Path"])
            selected_columns = ["Paraf", "Qty", "Path"]
            df = df[selected_columns]
            df = df._append(print_row,ignore_index=True)

        df = df.map(lambda x: str(x).replace(",", "."))
        df = df[df["Paraf"] != "nan"]
        if len(df) < 1:
            sys.exit()
        df["Qty"] = df["Qty"].astype(float)
        df["Qty"] = df["Qty"].astype(int)
        df = df.drop_duplicates(subset=["Paraf", "Qty", "Path"], keep="first")

        df = df.groupby(["Paraf"])["Qty"].sum().reset_index()
        if len(df) > 1:
            df = df.sort_values("Paraf")
        else:
            pass
        df["Code"] = df["Paraf"].apply(
            lambda x: (
                re.search(
                    r"(SERRMPR|ELECMPR|ELECTMC|ELEPAC|ELECMER|ELECFER)[^\s^\W]*", x
                ).group()
                if re.search(
                    r"(SERRMPR|ELECMPR|ELECTMC|ELEPAC|ELECMER|ELECFER)[^\s^\W]*", x
                )
                else ""
            )
        )
        df["Paraf"] = df["Paraf"].apply(
            lambda x: re.sub(
                r"(SERRMPR|ELECMPR|ELECTMC|ELEPAC|ELECMER|ELECFER)[^\s^\W]*", "", x
            )
        )

        df.insert(0, "Codigos", df.pop("Code"))

        df.columns = ["COD", "PAR", "QT"]
        df[["COD", "PAR", "QT"]] = df[["COD", "PAR", "QT"]]
        df_proj_append = pd.DataFrame()

        with open(
            options,
            "r",
        ) as file:
            lines = file.readlines()
        if lines:
            line = lines[9].strip()
            path_pro_data = (line.split("=")[1])[1:]
            path_pro = os.path.join(path_pro_data, r"Projeto\dados_proj.xlsx")
            df_proj = pd.read_excel(path_pro)

        with open(
            options,
            "r",
        ) as file:
            lines = file.readlines()
        if lines:
            line = lines[9].strip()
            path_pro_data = (line.split("=")[1])[1:]
            path_compras = os.path.join(path_pro_data, r"Compras\dados_compras.xlsx")
            df_compras = pd.read_excel(path_compras)

        for index, row in df.iterrows():
            if (row["COD"]) != "":
                pass
            if len(row["COD"]) > 9:  
                for (
                    index_proj,
                    row_proj,
                ) in (
                    df_proj.iterrows()
                ):  
                    code = (str(row["COD"]).lower()).strip()
                    code_proj = (str(row_proj["Codigo"]).lower()).strip()
                    par = ((row["PAR"]).lower()).strip()
                    par_proj = (row_proj["Design"]).lower().strip()
                    if code != code_proj and par == par_proj:
                        df_proj.at[index_proj, "Design"] = row[
                            "PAR"
                        ]  
                    elif (
                        code != code_proj
                    ):  
                        df_proj_append = (
                            pd.DataFrame()
                        )  
                        df_proj_append = df_proj_append._append(row)
                        df_proj_append = df_proj_append.drop(["QT"], axis=1)
                        df_proj_append = df_proj_append.rename(
                            columns={"COD": "Codigo", "PAR": "Design"}
                        )
                        df_proj = pd.concat([df_proj, df_proj_append])
                for index_compras, row_compras in df_compras.iterrows():
                    par_lower = row["PAR"].lower().strip()
                    design_lower = row_compras["Design"].lower().strip()
                    if design_lower != None:
                        if par_lower == design_lower:
                            df.at[index, "COD"] = row_compras["Codigo"]
            if str(row["COD"]) == "" or len(row["COD"]) < 9 or row["COD"] == None:
                for (
                    index_proj,
                    row_proj,
                ) in df_proj.iterrows():  
                    code = (str(row["COD"]).lower()).strip()
                    code_proj = (str(row_proj["Codigo"]).lower()).strip()
                    par = ((row["PAR"]).lower()).strip()
                    par_proj = (row_proj["Design"]).lower().strip()
                    if par_proj == par:
                        df.at[index, "COD"] = row_proj["Codigo"]
                for (
                    index_compras,
                    row_compras,
                ) in (
                    df_compras.iterrows()
                ):  
                    par_lower = row["PAR"].lower().strip()
                    design_lower = row_compras["Design"].lower().strip()
                    if par_lower == design_lower:
                        df.at[index, "COD"] = row_compras["Codigo"]

        df_proj["Design"] = (
            df_proj["Design"].str.strip().str.lower()
        )  
        df_proj = df_proj.drop_duplicates(keep="last")
        df_proj.to_excel(path_pro, index=False)

    except ValueError as e:
        error_message = f"Error: {str(e)}"
        problematic_row = df[df["Qty"].apply(lambda x: not str(x).isnumeric())]
        error_message += "\nLinha problematica:\n" + str(problematic_row)
        show_error_row(problematic_row, error_message)
        raise e

    df = df.sort_values(by=df.columns[1])
    workbook = Workbook()

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet = workbook.create_sheet(title="Material")
    sheet.insert_rows(0, 4)

    sheet["A1"] = ""
    sheet["A2"] = ""
    sheet["B1"] = ""
    title_sheet = "Lista Compra - " + file_name_1
    sheet["A3"] = " "
    sheet["A4"] = title_sheet
    sheet["A5"] = " "

    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    blue_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    very_light_brown_hex = "F5F5DC"
    very_light_brown_fill = PatternFill(
        start_color=very_light_brown_hex,
        end_color=very_light_brown_hex,
        fill_type="solid",
    )
    yellow_fill = PatternFill(
        start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    df = df.fillna(" ")
    dfE = df[df["COD"].str.startswith("ELE")]

    dfS = df[df["COD"].str.startswith(("SERR", " ", ""))]
    dfS = dfS[~dfS["COD"].str.startswith("ELEC")]

    for row in dataframe_to_rows(dfE, index=False, header=True):
        sheet.append(row)
    last_row_index = sheet.max_row

    sheet.move_range(f"A{last_row_index+1}:C{last_row_index+1}", rows=1)

    sheet.merge_cells(f"A{last_row_index+2}:C{last_row_index+2}")

    sheet[f"A{last_row_index+1}"] = "SERRALHARIA"

    merged_cell = sheet[f"A{last_row_index+1}"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
    merged_cell.border = thin_border
    merged_cell.fill = very_light_brown_fill
    merged_cell.font = Font(bold=True)


    for row in dataframe_to_rows(dfS, index=False, header=False):
        sheet.append(row)

    sheet.delete_cols(4)
    sheet["A6"] = "Codigo"
    sheet["B6"] = "Designação"
    sheet["C6"] = "Qt."

    for cell in sheet[1]:
        cell.font = header_font
    for cell in sheet[2]:
        cell.font = Font(bold=True)
    for cell in sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = blue_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for cell in sheet[6]:
        cell.font = Font(bold=True)
        cell.fill = gray_fill
    for row in sheet.iter_rows(min_row=6, min_col=1, max_row=sheet.max_row, max_col=3):
        for cell in row:
            cell.border = thin_border
    sheet.merge_cells("B1:C1")
    sheet.merge_cells("B2:C2")
    a1 = sheet["A1"]
    a2 = sheet["A2"]
    b1 = sheet["B1"]
    b2 = sheet["B2"]
    c1 = sheet["C1"]
    c2 = sheet["C2"]
    for cell in [a1, a2, b1, b2, c1, c2]:
        cell.border = border
    sheet.merge_cells("A4:C4")
    workbook.remove(workbook["Sheet"])

    sheet.insert_rows(7)
    a7 = sheet["A7"]
    b7 = sheet["B7"]
    c7 = sheet["C7"]
    for cell in [a7, b7, c7]:
        cell.border = border
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
    sheet.merge_cells("A7:C7")

    sheet["A7"] = "ELECTRICIDADE"
    merged_cell = sheet["A7"]
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

    if dfE.empty:
        sheet.insert_rows(8)
        a7 = sheet["A8"]
        b7 = sheet["B8"]
        c7 = sheet["C8"]
        for cell in [a7, b7, c7]:
            cell.border = border
            cell.font = Font(bold=True)
        sheet.merge_cells("A8:C8")
        sheet["A8"] = "Sem Componentes Elétricos"
        merged_cell = sheet["A8"]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("A9:C9")

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["C"].width = 5
    max_col_B_width = None
    max_paraf = max(len(str(value)) for value in df["PAR"])
    max_title = len(title_sheet)
    if max_paraf > max_title:
        max_col_B_width = max_paraf
    else:
        max_col_B_width = max_title

    column_B_letter = utils.get_column_letter(2)
    sheet.column_dimensions[column_B_letter].width = max_col_B_width
    workbook.save(output_file)



# SCRIPT START

with open(file_name, "r") as file:
    file_path = file.read().strip()

with open(source_name, "r") as file:
    source_path = file.read().strip()


data = pd.read_csv(file_path, delimiter="\t", encoding="latin-1")


file_path_txt = file_path
file_path = file_path[:-3] + "xlsx"
file_name = os.path.basename(file_path)

writer = pd.ExcelWriter(file_path, engine="xlsxwriter")
data.to_excel(writer, index=False, sheet_name="Sheet1")

writer._save()
file_name_1 = file_name[0:-5]
input_file = file_path  

parent_directory = os.path.dirname(source_path)
filename = os.path.basename(source_path)
filename_without_extension = os.path.splitext(filename)[0]
new_folder_name = "Listas_" + filename_without_extension

new_folder_path = os.path.join(parent_directory, new_folder_name)

with open(options, "r") as file:
    lines = file.readlines()

if lines:
    first_line = lines[
        0
    ].strip() 
    if first_line:
        macro_to_run = first_line[-1]

if not os.path.exists(new_folder_path):
    os.makedirs(new_folder_path)

default_file_name = new_folder_name + "_Excel"
file_name_final = (
    new_folder_path + "\\" + f"{filename_without_extension}_" + "Compras.xlsx"
)
output_file = file_name_final 

input_file = file_path
excel_to_dfs(input_file, output_file)
excel = win32com.client.Dispatch("Excel.Application")

workbook = excel.Workbooks.Open(output_file)

for worksheet in workbook.Worksheets:
    used_range = worksheet.UsedRange
    num_rows = used_range.Rows.Count
    num_cols = used_range.Columns.Count
    start_row = 3
    export_range = worksheet.Range(
        worksheet.Cells(start_row, 1), worksheet.Cells(num_rows, num_cols)
    )

    columns_range = export_range
    worksheet.PageSetup.PrintArea = columns_range.Address

    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = False

    pdf_export_format = 0  # 0 represents PDF format
    export_path = os.path.join(
        new_folder_path, f"{filename_without_extension}_Compras.pdf"
    )
    export_range.ExportAsFixedFormat(pdf_export_format, export_path)


workbook.Save()
workbook.Close()


excel.Quit()
logo_image_path = os.path.join(parent_dir, "extras", "Logo.pdf")
put_watermark(export_path, export_path, logo_image_path)
add_page_numbers(export_path)

window = tk.Tk()

window.withdraw()

if macro_to_run == "3":
    showMessage("Processo Terminado")
sys.exit()
