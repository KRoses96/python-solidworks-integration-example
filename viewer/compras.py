"""
Author: Manuel Rosa
This script creates a GUI to edit a dataframe and save it as a csv, all important data will be saved onto a database csv for future uses,
this also allows the user to find changes from the previous checked version and the newest version, in order to gain more control over each
version change.
"""

import os
import pandas as pd
import tkinter as tk
from tkinter import ttk, simpledialog, messagebox, Toplevel, Button
from tkinter import filedialog
from tkcalendar import DateEntry
from datetime import datetime
from tkcalendar import Calendar
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
import sys
import subprocess
from openpyxl import load_workbook
import shutil
import time
import threading
import math


pd.set_option("display.max_colwidth", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)


def nest_fix(df):
    """Cleans the nesting.csv"""
    df["Designação"] = (
        "Chapa " + df["Chapas"] + "x" + df["Esp."].astype(str) + "_" + df["Material"]
    )
    column_del = ["Material", "Esp.", "Chapas"]
    df = df.drop(columns=column_del)
    new_order = ["Codigo", "Designação", "Qt.", "Enc.", "Stock"]
    df = df[new_order]
    return df


def comp_fix(df):
    """Cleans the nesting.csv"""
    df["Enc."] = "0"
    df["Stock"] = "0"
    return df


def perfis_fix(df):
    """Cleans the perfis.csv"""
    df["Qt."] = df["Qt."].fillna(1)
    df["Comp.C.(m)"] = df["Comp.C.(m)"].str.extract(r"\((.*?)\)")
    df = df.fillna("")
    df["Designação"] = (
        df["Descrição"] + "_" + df["Material"].astype(str) + "-" + df["Comp.C.(m)"]
    )
    column_del = ["Descrição", "Material", "Comp.C.(m)"]
    df = df.drop(columns=column_del)
    new_order = ["Codigo", "Designação", "Comp.(m)", "Qt.", "Enc.", "Stock"]
    df = df[new_order]
    return df


script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
options = os.path.join(script_dir, "options.txt")
icon = os.path.join(script_dir, "md.ico")

with open(options, "r") as file:
    options = [line.strip() for line in file]

username = options[0].split("= ")[1]  
main_path = options[1].split("= ")[1]  
phc_path = options[2].split("= ")[
    1
]  
log_path = os.path.join(
    main_path, "log.txt"
)  


with open(log_path, "r") as file:
    lines = [line.strip() for line in file]
lines_proj = []
for line in lines:
    if line[0] == "V":  
        lines_proj.append(line)
    else:
        pass



all_versions = []
for versions in lines_proj:
    new_version = versions.split(":")[0]
    new_version_date = versions.split(",")[1]
    new_version_date = new_version_date.split(" ")[1]
    # Getting the right csv folder using version
    csv_folder = os.path.join(main_path, f"{new_version}")
    id_version = new_version + "-" + new_version_date
    if os.path.exists(csv_folder):
        all_versions.append(id_version)

version_dict = {}  
for date in all_versions:
    prefix = date.split("-")[0]
    version_dict[prefix] = date

all_versions = list(version_dict.values())


i = 1
for version in lines_proj:
    last_version_pro = lines_proj[-i]
    last_v = last_version_pro.split(":")[0]
    author_pro = ((last_version_pro.split(":")[1]).split(",")[0])[1:]
    csv_folder = os.path.join(main_path, f"{last_v}")
    csv_path = os.path.join(csv_folder, "CSV")
    if os.path.exists(csv_path):
        break
    else:
        i = i + 1

control_path = os.path.join(csv_path, "controlo.csv")
pro_name = csv_folder.split("\\")

project_name = (
    pro_name[-3] + "-" + pro_name[-2] + "_" + pro_name[-1] + "(" + author_pro + ")"
)

lines_review = []
last_reviewed_ver = None
repeat_lines = None
for line in lines:
    if line[0] == "_":
        lines_review.append(line)
    else:
        pass


latest_timestamps = {}

for entry in lines_review:
    version, timestamp = entry.split(":")[0], entry.split(",")[1]
    if version in latest_timestamps:
        if timestamp > latest_timestamps[version]:
            latest_timestamps[version] = timestamp
    else:
        latest_timestamps[version] = timestamp

lines_review = [
    "{}: {}".format(version, timestamp)
    for version, timestamp in latest_timestamps.items()
]
lines_review.sort()


if lines_review != []:
    last_reviewed_ver_line = lines_review[-1]
    last_reviewed_ver = (last_reviewed_ver_line.split(":")[0])[1:]
    if last_reviewed_ver == last_v and len(lines_review) > 1:
        last_reviewed_ver_line = lines_review[-2]
        last_reviewed_ver = (last_reviewed_ver_line.split(":")[0])[1:]
    csv_folder_review = os.path.join(main_path, f"{last_reviewed_ver}")
    csv_path_review = os.path.join(csv_folder_review, "CSV")
    control_path_review = os.path.join(csv_path_review, "controlo.csv")


def not_checked_df(csv_path):
    """This cleans dataframes that haven't been reviewed and need to be processed.
    If the version does not have controlo.csv"""
    compras_csv = os.path.join(csv_path, "Compras.csv")
    dxf_csv = os.path.join(csv_path, "DXF.csv")
    nest_csv = os.path.join(csv_path, "NEST.csv")
    perfis_csv = os.path.join(csv_path, "Perfis.csv")
    if os.path.exists(compras_csv):
        df_comp = pd.read_csv(compras_csv)
        df_comp = comp_fix(df_comp)
    else:
        df_comp = pd.DataFrame()
    if os.path.exists(nest_csv):
        df_nest = pd.read_csv(nest_csv)
        df_nest = nest_fix(df_nest)
    else:
        df_nest = pd.DataFrame()
    if os.path.exists(perfis_csv):
        df_perfis = pd.read_csv(perfis_csv)
        df_perfis = perfis_fix(df_perfis)
    else:
        df_perfis = pd.DataFrame()

    df_global = pd.concat([df_nest, df_comp, df_perfis])
    df_global.dropna(subset=["Designação"], inplace=True)
    df_global = df_global.sort_values(by="Designação")

    df_global["Qt."] = df_global["Qt."].astype(int)
    new_global_order = [
        "Codigo",
        "Designação",
        "Comp.(m)",
        "Qt.",
        "Stock",
        "Enc.",
    ]  

    if (
        "Comp.(m)" not in df_global.columns
    ):
        df_global["Comp.(m)"] = ""

    df_global = df_global[new_global_order]
    df_global["Data Enc."] = ""
    df_global["Entr."] = "0"
    df_global = df_global.fillna("")

    if last_reviewed_ver != None:
        df_global_old = pd.read_csv(control_path_review)
        df_global_old = df_global_old.fillna("")
        df_global.set_index("Designação", inplace=True)
        df_global_old.set_index("Designação", inplace=True)

        columns_to_update = ["Stock", "Enc.", "Data Enc.", "Entr."]
        df_global[columns_to_update] = df_global_old[columns_to_update].combine_first(
            df_global[columns_to_update]
        )

        df_global = df_global.fillna("")

        df_global.reset_index(inplace=True)

        new_global_order = [
            "Codigo",
            "Designação",
            "Comp.(m)",
            "Qt.",
            "Stock",
            "Enc.",
            "Data Enc.",
            "Entr.",
        ]
        df_global = df_global[new_global_order]

    return df_global


if os.path.exists(control_path):
    df_global = pd.read_csv(control_path)
    df_global = df_global.fillna("")
    df_global["Stock"] = df_global["Stock"].astype(str)
    df_global["Stock"] = df_global["Stock"].replace("0", "☐")
    df_global["Stock"] = df_global["Stock"].replace("1", "☒")
    df_global["Enc."] = df_global["Enc."].astype(str)
    df_global["Enc."] = df_global["Enc."].replace("0", "☐")
    df_global["Enc."] = df_global["Enc."].replace("1", "☒")
    df_global["Entr."] = df_global["Entr."].astype(str)
    df_global["Entr."] = df_global["Entr."].replace("0", "☐")
    df_global["Entr."] = df_global["Entr."].replace("1", "☒")
else:
    df_global = not_checked_df(csv_path)
    df_global["Stock"] = df_global["Stock"].astype(int)
    df_global["Stock"] = df_global["Stock"].astype(str)
    df_global["Stock"] = df_global["Stock"].replace("0", "☐")
    df_global["Stock"] = df_global["Stock"].replace("1", "☒")
    df_global["Enc."] = df_global["Enc."].astype(int)
    df_global["Enc."] = df_global["Enc."].astype(str)
    df_global["Enc."] = df_global["Enc."].replace("0", "☐")
    df_global["Enc."] = df_global["Enc."].replace("1", "☒")
    df_global["Entr."] = df_global["Entr."].astype(int)
    df_global["Entr."] = df_global["Entr."].astype(str)
    df_global["Entr."] = df_global["Entr."].replace("0", "☐")
    df_global["Entr."] = df_global["Entr."].replace("1", "☒")
    df_global = df_global.reset_index(drop=True)

comp_excel = os.path.join(phc_path, "Compras")
comp_excel = os.path.join(
    comp_excel, "dados_compras.xlsx"
)  

phc_excel = os.path.join(phc_path, "phc.xlsx")

df_comp = pd.read_excel(comp_excel)
df_comp.rename(columns={"Design": "Designação"}, inplace=True)


for index, row in df_global.iterrows():
    for index_comp, row_comp in df_comp.iterrows():
        desig_comp = row_comp["Designação"]
        desig_normal = row["Designação"]
        code_comp = row_comp["Codigo"]
        if desig_normal.lower().strip() == desig_comp.lower().strip():
            df_global.at[index, "Codigo"] = code_comp


add_path = os.path.join(main_path, "add.csv")
if os.path.exists(add_path):
    df_added = pd.read_csv(add_path)

    for index, row in df_added.iterrows():
        desig_row = row["Designação"]
        for index_real, row_real in df_global.iterrows():
            if row_real["Designação"] == desig_row:
                df_global = df_global.drop(index_real)

    df_added.fillna("", inplace=True)
    df_added["Stock"] = df_added["Stock"].replace(0, "☐")
    df_added["Stock"] = df_added["Stock"].replace(1, "☒")
    df_added["Enc."] = df_added["Enc."].replace(0, "☐")
    df_added["Enc."] = df_added["Enc."].replace(1, "☒")
    df_added["Entr."] = df_added["Entr."].replace(0, "☐")
    df_added["Entr."] = df_added["Entr."].replace(1, "☒")
    list_added = []
    df_added_appends = pd.DataFrame()
    for index, row in df_global.iterrows():
        list_added.append(row["Designação"])

    for index, row in df_added.iterrows():
        if row["Designação"] not in list_added:
            df_added_appends = df_added_appends._append(row)
    df_global = df_global._append(df_added_appends, ignore_index=True)


def compare_data(df_current, version_old):
    """
    Function used to compare between two dataframes
    df_current:Dataframe from the newest version of the project
    version_old: Dataframe from any of the older versions of the project
    """

    df_current["Stock"] = df_current["Stock"].astype(str)
    df_current["Stock"] = df_current["Stock"].replace("☐", "0")
    df_current["Stock"] = df_current["Stock"].replace("☒", "1")
    df_current["Enc."] = df_current["Enc."].astype(str)
    df_current["Enc."] = df_current["Enc."].replace("☐", "0")
    df_current["Enc."] = df_current["Enc."].replace("☒", "1")

    csv_folder_old = os.path.join(main_path, f"{version_old}")
    csv_path_old = os.path.join(csv_folder_old, "CSV")
    control_path_old = os.path.join(csv_path_old, "controlo.csv")

    if os.path.exists(control_path_old):
        df_old = pd.read_csv(control_path_old)
    else:
        compras_csv_old = os.path.join(csv_path_old, "Compras.csv")
        dxf_csv_old = os.path.join(
            csv_path_old, "DXF.csv"
        ) 
        nest_csv_old = os.path.join(csv_path_old, "NEST.csv")
        perfis_csv_old = os.path.join(csv_path_old, "Perfis.csv")

        if os.path.exists(compras_csv_old):
            df_comp_old = pd.read_csv(compras_csv_old)
            df_comp_old = comp_fix(df_comp_old)
        else:
            df_comp_old = pd.DataFrame()
        if os.path.exists(nest_csv_old):
            df_nest_old = pd.read_csv(nest_csv_old)
            df_nest_old = nest_fix(df_nest_old)
        else:
            df_nest_old = pd.DataFrame()
        if os.path.exists(perfis_csv_old):
            df_perfis_old = pd.read_csv(perfis_csv_old)
            df_perfis_old = perfis_fix(df_perfis_old)
        else:
            df_perfis_old = pd.DataFrame()

        df_old = pd.concat([df_nest_old, df_comp_old, df_perfis_old])
        df_old = df_old.sort_values(by="Designação")
        df_old["Qt."] = df_old["Qt."].astype(int)
        if "Comp.(m)" not in df_old.columns:
            df_old["Comp.(m)"] = ""
        new_global_order = ["Codigo", "Designação", "Comp.(m)", "Qt.", "Stock", "Enc."]

        df_old = df_old[new_global_order]
        df_old["Data Enc."] = ""
        df_old["Entr."] = ""

    df_old = df_old.fillna("")

    merged_df = pd.merge(
        df_current, df_old, on="Designação", how="outer", suffixes=("_main", "_old")
    )

    diff_rows = merged_df[
        (merged_df["Qt._main"] != merged_df["Qt._old"])
        | (merged_df["Comp.(m)_main"] != merged_df["Comp.(m)_old"])
    ]
    diff_rows_cleaned = diff_rows.dropna()

    new_designacoes = merged_df[merged_df["Qt._old"].isnull()]

    removed_designacoes = merged_df[merged_df["Qt._main"].isnull()]

    designacao_changed = diff_rows_cleaned["Designação"].tolist()
    designacao_added = new_designacoes["Designação"].tolist()
    designacao_removed = removed_designacoes["Designação"].tolist()

    df_removed = removed_designacoes[
        [
            "Codigo_old",
            "Designação",
            "Qt._old",
            "Comp.(m)_old",
            "Stock_old",
            "Enc._old",
            "Data Enc._old",
            "Entr._old",
        ]
    ]
    df_removed.columns = [
        "Codigo",
        "Designação",
        "Qt.",
        "Comp.(m)",
        "Stock",
        "Enc.",
        "Data Enc.",
        "Entr.",
    ]
    df_current["Qt."] = df_current["Qt."].astype(int)

    # Replace the 0 and 1 back into checkboxes for GUI - quickfix to meet the deadline
    df_current["Stock"] = df_current["Stock"].astype(str)
    df_current["Stock"] = df_current["Stock"].replace("0", "☐")
    df_current["Stock"] = df_current["Stock"].replace("1", "☒")
    df_current["Stock"] = df_current["Stock"].replace("0.0", "☐")
    df_current["Stock"] = df_current["Stock"].replace("1.0", "☒")

    df_current["Enc."] = df_current["Enc."].astype(str)
    df_current["Enc."] = df_current["Enc."].replace("0", "☐")
    df_current["Enc."] = df_current["Enc."].replace("1", "☒")
    df_current["Enc."] = df_current["Enc."].replace("0.0", "☐")
    df_current["Enc."] = df_current["Enc."].replace("1.0", "☒")

    df_current["Entr."] = df_current["Entr."].astype(str)
    df_current["Entr."] = df_current["Entr."].replace("0", "☐")
    df_current["Entr."] = df_current["Entr."].replace("1", "☒")
    df_current["Entr."] = df_current["Entr."].replace("0.0", "☐")
    df_current["Entr."] = df_current["Entr."].replace("1.0", "☒")

    df_removed.loc[:, "Qt."] = df_removed["Qt."].astype(str).str.rstrip(".0")
    df_removed.loc[:, "Enc."] = df_removed["Enc."].astype(str)
    df_removed.loc[:, "Enc."] = df_removed["Enc."].replace("1.0", "☒")
    df_removed.loc[:, "Enc."] = df_removed["Enc."].replace("0.0", "☐")
    df_removed.loc[:, "Enc."] = df_removed["Enc."].replace("1", "☒")
    df_removed.loc[:, "Enc."] = df_removed["Enc."].replace("0", "☐")

    df_removed.loc[:, "Stock"] = df_removed["Stock"].astype(int)
    df_removed.loc[:, "Stock"] = df_removed["Stock"].astype(str)
    df_removed.loc[:, "Stock"] = df_removed["Stock"].replace("1.0", "☒")
    df_removed.loc[:, "Stock"] = df_removed["Stock"].replace("0.0", "☐")
    df_removed.loc[:, "Stock"] = df_removed["Stock"].replace("1", "☒")
    df_removed.loc[:, "Stock"] = df_removed["Stock"].replace("0", "☐")

    return designacao_changed, designacao_added, designacao_removed, df_removed


class DataFrameEditor(tk.Tk):
    def __init__(self, dataframe):
        super().__init__()

        self.title(project_name)
        self.dataframe = dataframe
        self.real_dataframe = (
            dataframe  
        )
        self.filtered_dataframe = dataframe 
        self.compare_dataframe = (
            pd.DataFrame()
        )  
        self.compare = tk.BooleanVar(value=False)  
        self.removed = tk.BooleanVar(value=False)  
        self.edited = tk.BooleanVar(value=False)  
        self.added = tk.BooleanVar(value=False)  
        self.dropdown_var = None  
        self.searching_bol = False  
        self.search_var = None
        self.versions = all_versions[
            :-1
        ]  
        self.desig_value = (
            None 
        )
        self.item_selected = (
            None
        )
        self.add_counter = 0 
        self.df_comp = pd.read_excel(
            comp_excel
        )  
        self.df_comp.rename(
            columns={"Design": "Designação"}, inplace=True
        )  
        self.new_add = 1  
        self.last_loaded_df = None
        self.df_added = (
            pd.DataFrame()
        )  
        self.df_phc = pd.read_excel(
            phc_excel
        ) 
        self.new_value = None  
        self.update_dialog = (
            0  # Used to force update the new_value from Toplevel windows
        )
        self.create_widgets()

    def create_widgets(self):
        """Creates all widgets for the main window"""
        frame1 = ttk.Frame(self)
        frame1.grid(row=0, column=0, padx=0, pady=10, sticky="nsew")
        scrollbar = ttk.Scrollbar(frame1, orient="vertical")
        scrollbar.grid(row=0, column=1, sticky="nsew")

        self.treeview = ttk.Treeview(
            frame1,
            columns=list(self.dataframe.columns),
            show="headings",
            yscrollcommand=scrollbar.set,
            height=24,
        )

        scrollbar.config(command=self.treeview.yview)
        scrollbar.bind(
            "<B1-Motion>",
            lambda event: self.treeview.yview_scroll(
                int(-1 * (event.delta / 120)), "units"
            ),
        )

        scrollbar.bind(
            "<MouseWheel>",
            lambda event: self.treeview.yview_scroll(
                int(-1 * (event.delta / 120)), "units"
            ),
        )

        for col in self.dataframe.columns:
            self.treeview.heading(col, text=col)
        for col in self.dataframe.columns:
            max_width = max(self.dataframe[col].astype(str).apply(lambda x: len(x))) + 4
            padding = 4  
            self.treeview.column(col, width=(max_width + padding) * 7)
            if col == "Qt.":
                self.treeview.column(col, width=60)
            if col == "Stock":
                self.treeview.column(col, width=45)
            if col == "Entr.":
                self.treeview.column(col, width=45)
            if col == "Enc.":
                self.treeview.column(col, width=45)
            if col == "Data Enc.":
                self.treeview.column(col, width=90)
            if col == "Comp.(m)":
                self.treeview.column(col, width=80)
            if col == "Codigo":
                self.treeview.column(col, width=100)

        all_columns = [""]
        for col in self.dataframe.columns:
            if (
                col == "Qt."
                or col == "Enc."
                or col == "Stock"
                or col == "Data Enc."
                or col == "Comp.(m)"
                or col == "Entr."
            ):
                self.treeview.heading(col, text=col, anchor=tk.CENTER)
                self.treeview.column(col, anchor=tk.CENTER)
            all_columns.append(col)

        self.treeview.bind("<ButtonRelease-1>", self.select_item)
        self.treeview.bind("<Double-1>", self.edit_cell)

        self.treeview.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        frame2 = ttk.Frame(self)
        frame2.grid(row=0, column=1, padx=10, pady=5, sticky="nsew", rowspan=2)

        underline_font = ("TkDefaultFont", 11, "underline")
        search_label = ttk.Label(frame2, text="Procurar:", font=underline_font)
        search_label.grid(row=0, column=0, padx=(10, 5), pady=5)

        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(frame2, textvariable=self.search_var)
        search_entry.grid(
            row=0, column=1, columnspan=2, padx=(0, 5), pady=5, sticky="we"
        )
        self.search_var.trace_add("write", self.search_bol_update)

        separator_1 = ttk.Separator(frame2, orient="horizontal")
        separator_1.grid(row=1, column=0, columnspan=3, sticky="ew", pady=5)

        options = self.versions
        option_on_startup = last_reviewed_ver
        option_on_start = None
        if options != []: 
            for option in options:
                if option_on_startup:
                    if option.startswith(option_on_startup):
                        option_on_start = option
                    else:
                        pass
            if option_on_start == None:
                option_on_start = options[-1]
            else:
                pass

        self.dropdown_var = tk.StringVar(value=option_on_start)  
        self.dropdown_menu = ttk.Combobox(
            frame2, values=options, state="readonly", textvariable=self.dropdown_var
        )
        self.dropdown_menu.grid(row=2, column=1, padx=5, pady=5)
        self.dropdown_menu.bind("<<ComboboxSelected>>", self.on_combobox_change)

        self.hist_button = ttk.Button(frame2, text="Histórico", command=self.history)
        self.hist_button.grid(row=2, column=2, padx=5, pady=5)

        self.comp_button = ttk.Checkbutton(
            frame2,
            text=" Comparar ",
            variable=self.compare,
            command=self.update_comp_button,
        )
        self.removed_button = ttk.Checkbutton(
            frame2,
            text="Removidos",
            variable=self.removed,
            command=self.update_remo_button,
        )
        self.edited_button = ttk.Checkbutton(
            frame2,
            text="Editados",
            variable=self.edited,
            command=self.update_edit_button,
        )
        self.added_button = ttk.Checkbutton(
            frame2,
            text="Adicionados",
            variable=self.added,
            command=self.update_add_button,
        )

        if (
            options == []
        ):  
            self.comp_button["state"] = "disabled"
            self.hist_button["state"] = "disabled"

        self.comp_button.grid(row=2, column=0, padx=5, pady=5)
        self.removed_button.grid(row=3, column=0, padx=5, pady=5)
        self.edited_button.grid(row=3, column=1, padx=5, pady=5)
        self.added_button.grid(row=3, column=2, padx=5, pady=5)


        separator_2 = ttk.Separator(frame2, orient="horizontal")
        separator_2.grid(row=4, column=0, columnspan=3, sticky="ew", pady=5)


        save_button = ttk.Button(frame2, text="Guardar", command=self.save_csv)
        save_button.grid(row=5, column=0, padx=5, pady=5)

        export_button = ttk.Button(
            frame2, text="Excel", command=self.export_dataframe_to_excel
        )
        export_button.grid(row=5, column=1, padx=5, pady=5)


        add_button = ttk.Button(frame2, text="Adicionar", command=self.add_to_data)
        add_button.grid(row=5, column=2, padx=5, pady=5)


        separator_3 = ttk.Separator(frame2, orient="horizontal")
        separator_3.grid(row=6, column=0, columnspan=3, sticky="ew", pady=5)

        note_label = ttk.Label(frame2, text="Anotações", font=underline_font)
        note_label.grid(row=7, column=1, padx=(10, 5), pady=5)

        self.note_text = tk.Text(frame2, wrap="word", height=19, width=2)
        self.note_text.grid(
            row=8, column=0, columnspan=3, padx=(0, 5), pady=5, sticky="nsew"
        )
        self.note_text.config(font=("TkDefaultFont", 11))

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        frame2.columnconfigure(0, weight=1)
        frame2.columnconfigure(2, weight=1)
        self.resizable(
            False, False
        ) 

        self.update_idletasks()
        self.update_treeview()
        self.txt_read()
        self.geometry(f"{self.winfo_reqwidth()+70}x600")

    def select_item(self, event):
        """
        The function selects an item from a treeview widget and retrieves information about the selected
        item.
        """
        if self.treeview.selection():
            self.item_selected = self.treeview.selection()[0]
            self.desig_value = self.treeview.item(self.item_selected, "values")[1]
        else:
            self.item_selected = None

    def history_fetcher(self, desig):
        """
        fetches history from the selected item and creates a dataframe with that information
        """

        all_version_paths = []
        df_history = pd.DataFrame()
        for ver in all_versions:
            version_folder = ver.split("-")[0]
            csv_v_folder = os.path.join(main_path, version_folder)
            all_version_paths.append(os.path.join(csv_v_folder, "CSV"))
        for version_path in all_version_paths:
            version_p = os.path.dirname(version_path)
            version_name = os.path.basename(version_p)
            v_control = os.path.join(version_path, "controlo.csv")
            if os.path.exists(v_control):
                df_control = pd.read_csv(v_control)
                for index, row in df_control.iterrows():
                    if desig in row["Designação"]:
                        row["Versão"] = version_name
                        row["Verif."] = "✓"
                        df_history = df_history._append(row)
            else:
                df_ver = not_checked_df(version_path)
                for index, row in df_ver.iterrows():
                    if desig in row["Designação"]:
                        row["Versão"] = version_name
                        row["Verif."] = ""
                        df_history = df_history._append(row)
        
        df_history = df_history.fillna("")

        column_order = [
            "Versão",
            "Codigo",
            "Designação",
            "Comp.(m)",
            "Qt.",
            "Stock",
            "Enc.",
            "Data Enc.",
            "Verif.",
        ]

        df_history = df_history[column_order]

        df_history["Stock"] = df_history["Stock"].astype(int)
        df_history["Stock"] = df_history["Stock"].astype(float)

        df_history["Stock"] = df_history["Stock"].replace(1.0, "☒")
        df_history["Stock"] = df_history["Stock"].replace(0.0, "☐")
        df_history["Stock"] = df_history["Stock"].replace("1", "☒")
        df_history["Stock"] = df_history["Stock"].replace("0", "☐")

        df_history["Enc."] = df_history["Enc."].replace(1.0, "☒")
        df_history["Enc."] = df_history["Enc."].replace(0.0, "☐")
        df_history["Enc."] = df_history["Enc."].replace("0", "☐")
        df_history["Enc."] = df_history["Enc."].replace("1", "☒")

        return df_history

    def history(self):
        """Opens a new window of the row selected showing the history of the item
        over the past versions of the project"""

        selected_item = self.desig_value
        df_history = self.history_fetcher(selected_item)

        if selected_item and not df_history.empty:
            top = tk.Toplevel()
            top.title(f"Histórico: {selected_item}")
            top.iconbitmap(icon)


            tree = ttk.Treeview(top)
            tree["columns"] = tuple(df_history.columns)
            tree["show"] = "headings"

            for col in df_history.columns:
                tree.heading(col, text=col)
            for col in df_history.columns:
                max_width = max(df_history[col].astype(str).apply(lambda x: len(x))) + 4
                padding = 4  
                tree.column(col, width=(max_width + padding) * 7)
                if col == "Qt.":
                    tree.column(col, width=40)
                if col == "Stock":
                    tree.column(col, width=45)
                if col == "Verif.":
                    tree.column(col, width=45)
                if col == "Enc.":
                    tree.column(col, width=35)
                if col == "Data Enc.":
                    tree.column(col, width=90)
                if col == "Comp.(m)":
                    tree.column(col, width=80)
                if col == "Codigo":
                    tree.column(col, width=100)

            all_columns = [""]
            for col in df_history.columns:
                if (
                    col == "Qt."
                    or col == "Enc."
                    or col == "Stock"
                    or col == "Data Enc."
                    or col == "Comp.(m)"
                    or col == "Verif."
                ):
                    tree.heading(col, text=col, anchor=tk.CENTER)
                    tree.column(col, anchor=tk.CENTER)
                all_columns.append(col)

            for _, row in df_history.iterrows():
                tree.insert("", "end", values=tuple(row))
            scrollbar = ttk.Scrollbar(top, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            tree.pack(fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

        else:
            pass

    def txt_read(self):
        """
        Saves the user input from the GUI to the txt inside the folder of the project
        """
        txt_name = "notas.txt"
        txt_path = os.path.join(main_path, txt_name)
        if os.path.exists(txt_path):
            with open(txt_path, "r", encoding="utf-8") as file:
                file_content = file.read()
                self.note_text.delete(1.0, tk.END)  # Clear existing text
                self.note_text.insert(tk.END, file_content)

    def add_to_data(self):
        """
        Opens a window that allows the user to add extra elements to the original treeview
        """
        self.add_counter = self.add_counter + 1  # Counter for Threads
        internal_counter = (
            self.add_counter
        )  
        add_window = tk.Toplevel()
        add_window.title("Adição")
        add_window.iconbitmap(icon)
        add_window.resizable(False, False)

        df_add_current = pd.DataFrame(columns=["Codigo", "Designação", "Qt."])

        path_add_excel = os.path.join(main_path, "add.xlsx")
        path_add_csv = os.path.join(main_path, "add.csv")

        if os.path.exists(
            path_add_excel
        ):  
            df_add_current = pd.read_excel(path_add_excel)

        def update_tree():
            first_loop = 1  
            while True:
                if internal_counter != self.add_counter:
                    break  
                if os.path.exists(path_add_excel):
                    df_add_current = pd.read_excel(path_add_excel)
                    if len(df_add_current) > 0:
                        if (
                            self.last_loaded_df is None
                            or not df_add_current.equals(self.last_loaded_df)
                            or first_loop == 1
                        ):
                            self.last_loaded_df = df_add_current
                            for item in treeview_a.get_children():  
                                treeview_a.delete(item)
                            for (
                                index,
                                row,
                            ) in (
                                df_add_current.iterrows()
                            ): 
                                values = list(row)
                                values = [
                                    (
                                        ""
                                        if (isinstance(x, float) and math.isnan(x))
                                        else int(x) if i == 2 else x
                                    )
                                    for i, x in enumerate(values)
                                ]  
                                if values[1] != "" or values[2] != "":
                                    treeview_a.insert(
                                        "", "end", text=index, values=values
                                    )
                else:
                    df_add_current = pd.DataFrame(
                        columns=["Codigo", "Designação", "Qt."]
                    )
                first_loop = 0
                time.sleep(5)

        update_t = threading.Thread(target=update_tree)
        update_t.daemon = True
        update_t.start()

        frame_tree_a = ttk.Frame(add_window)
        frame_tree_a.grid(row=0, column=0, padx=5, pady=5, sticky="nswe")

        frame_buttons = ttk.Frame(add_window)
        frame_buttons.grid(row=1, column=0, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(frame_tree_a, orient="vertical")
        scrollbar.grid(row=1, column=1, sticky="nsew")

        treeview_a = ttk.Treeview(
            frame_tree_a,
            columns=list(df_add_current.columns),
            show="headings",
            yscrollcommand=scrollbar.set,
        )

        scrollbar.config(command=treeview_a.yview)
        treeview_a.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        for col in df_add_current.columns:
            treeview_a.heading(col, text=col)
        for col in list(df_add_current.columns):
            if col == "Codigo":
                treeview_a.column(col, width=100)
            if col == "Designação":
                treeview_a.column(col, width=500)
            if col == "Qt.":
                treeview_a.column(col, width=60)
                treeview_a.heading(col, text=col, anchor=tk.CENTER)
                treeview_a.column(col, anchor=tk.CENTER)
        for index, row in df_add_current.iterrows():
            values = list(row)
            treeview_a.insert("", "end", text=index, values=values)

        def new_addition(open_excel=1):
            """
            Used to create the excel that feeds the GUI
            open_excel = 1 opens the excel automatically, open_excel = 0 does not open the excel
            """
            self.new_add = 1

            if os.path.exists(path_add_excel):
                self.new_add = 0  
                confirmation_box = tk.Toplevel(root)
                confirmation_box.title("Alerta")
                confirmation_box.iconbitmap(icon)
                confirmation_box.resizable(0, 0)
                confirmation_box.attributes("-topmost", "true")

                def yes_del():
                    self.new_add = 1
                    confirmation_box.destroy()

                def no_del():
                    confirmation_box.destroy()

                info_yesno = ttk.Label(
                    confirmation_box, text="Deseja apagar a lista atual?"
                )
                info_yesno.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

                yes_button = ttk.Button(confirmation_box, text="Sim", command=yes_del)
                yes_button.grid(row=1, column=0, padx=5, pady=5)

                no_button = ttk.Button(confirmation_box, text="Não", command=no_del)
                no_button.grid(row=1, column=1, padx=5, pady=5)

                confirmation_box.wait_window()
            if self.new_add == 1:
                try:
                    with pd.ExcelWriter(path_add_excel, engine="xlsxwriter") as writer:

                        df_add = pd.DataFrame(columns=["Codigo", "Designação", "Qt."])
                        df_add.to_excel(writer, sheet_name="Sheet1", index=False)

                        workbook = writer.book
                        worksheet = writer.sheets["Sheet1"]

                        df_add.to_excel(writer, sheet_name="Sheet1", index=False)

                        workbook = writer.book
                        worksheet = writer.sheets["Sheet1"]

                        header_format = workbook.add_format(
                            {
                                "bold": True,
                                "text_wrap": True,
                                "valign": "top",
                                "fg_color": "#004080",  
                                "color": "white",
                                "align": "center",
                                "border": 1,
                            }
                        )
                        for col_num, value in enumerate(df_add.columns.values):
                            worksheet.write(0, col_num, value, header_format)


                        worksheet.set_column("B:B", 50)
                        worksheet.set_column("A:A", 15)
                        error = 0
                except:
                    messagebox.showerror(
                        "Erro", "Excel aberto, porfavor feche antes de continuar"
                    )
                    add_window.destroy()
                    error = 1
                if error == 0:  
                    if open_excel == 1:
                        subprocess.Popen(path_add_excel, shell=True)

        def edit_addition():
            """Opens the add.xlsx to be edited by the user"""
            if os.path.exists(path_add_excel):
                try:
                    os.startfile(path_add_excel)
                except:
                    messagebox.showerror(
                        "Erro",
                        "Excel já se encontra aberto em outro computador,\nVerifique com os outros utilizadores!",
                    )
            else:
                new_addition()

        def import_addition():
            """Imports rows inside an xlsx into the xlsx inside the project folder"""
            template = filedialog.askopenfilename(
                defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
            )
            df_template = pd.read_excel(template)
            df_current = pd.read_excel(path_add_excel)
            df_final = pd.concat([df_template, df_current], ignore_index=True)

            df_final.drop_duplicates(inplace=True)

            correct_cols = ["Codigo", "Designação", "Qt."]
            if df_template.columns.tolist() == correct_cols:
                try:
                    with pd.ExcelWriter(path_add_excel, engine="xlsxwriter") as writer:
                        print(len(df_final))
                        if len(df_final) < 1:
                            df_add = pd.DataFrame(
                                columns=["Codigo", "Designação", "Qt."]
                            )
                        else:
                            df_add = df_final
                        df_add.to_excel(writer, sheet_name="Sheet1", index=False)

                        workbook = writer.book
                        worksheet = writer.sheets["Sheet1"]

                        df_add.to_excel(writer, sheet_name="Sheet1", index=False)

                        workbook = writer.book
                        worksheet = writer.sheets["Sheet1"]

                        header_format = workbook.add_format(
                            {
                                "bold": True,
                                "text_wrap": True,
                                "valign": "top",
                                "fg_color": "#004080",  
                                "color": "white",
                                "align": "center",
                                "border": 1,
                            }
                        )
                        for col_num, value in enumerate(df_add.columns.values):
                            worksheet.write(0, col_num, value, header_format)

                        worksheet.set_column("B:B", 50)
                        worksheet.set_column("A:A", 15)

                        error = 0
                        open_file = open(path_add_excel, "r+")
                except:
                    messagebox.showerror(
                        "Erro", "Excel aberto, porfavor feche antes de continuar"
                    )
                    add_window.destroy()
                    error = 1
            else:
                messagebox.showerror("Erro", "Template inválido!")
            add_window.lift()  

        def export_addition():
            """Moves the xlsx inside the project folder somewhere else"""
            export_temp_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
            )
            shutil.copy(path_add_excel, export_temp_path)
            add_window.lift()  

        def add_addition():
            """
            Saves the df_added to the main treeview
            """
            if os.path.exists(path_add_excel):
                df_added = pd.read_excel(path_add_excel)
                df_added["Comp.(m)"] = ""
                df_added["Stock"] = "0"
                df_added["Enc."] = "0"
                df_added["Data Enc."] = ""
                df_added["Entr."] = "0"
                df_added = df_added[
                    [
                        "Codigo",
                        "Designação",
                        "Comp.(m)",
                        "Qt.",
                        "Stock",
                        "Enc.",
                        "Data Enc.",
                        "Entr.",
                    ]
                ]

                # In case there was already a previous addition to the project,
                # save information from the columns : ´Stock´,´Enc.´,´Data Enc.´,´Entr.´
                # from old items still present in this addition
                if os.path.exists(path_add_csv):
                    df_added_old = pd.read_csv(path_add_csv)
                    for index, row in df_added_old.iterrows():
                        desig_row = row["Designação"]
                        for index_real, row_real in self.real_dataframe.iterrows():
                            if row_real["Designação"] == desig_row:
                                self.real_dataframe = self.real_dataframe.drop(
                                    index_real
                                )
                        for index_new, row_new in df_added.iterrows():
                            if row_new["Designação"] == row["Designação"]:
                                row_new["Stock."] = row["Stock"]
                                row_new["Enc."] = row["Enc."]
                                row_new["Entr."] = row["Entr."]
                                row_new["Data Enc."] = row["Data Enc."]
                df_added.dropna(subset=["Designação"], inplace=True)
                df_added.fillna("", inplace=True)
                df_added.to_csv(
                    path_add_csv, index=False
                )  

                list_added = []
                df_added_appends = pd.DataFrame()
                for (
                    index,
                    row,
                ) in (
                    self.real_dataframe.iterrows()
                ):  
                    list_added.append(row["Designação"])

                for (
                    index,
                    row,
                ) in (
                    df_added.iterrows()
                ): 
                    if row["Designação"] not in list_added:
                        df_added_appends = df_added_appends._append(row)

                # Quickfix of values after being appended to the permanent dataframe
                self.real_dataframe = self.real_dataframe._append(
                    df_added_appends, ignore_index=True
                )
                self.real_dataframe["Stock"] = self.real_dataframe["Stock"].replace(
                    0, "☐"
                )
                self.real_dataframe["Stock"] = self.real_dataframe["Stock"].replace(
                    1, "☒"
                )
                self.real_dataframe["Stock"] = self.real_dataframe["Stock"].replace(
                    "0", "☐"
                )
                self.real_dataframe["Stock"] = self.real_dataframe["Stock"].replace(
                    "1", "☒"
                )
                self.real_dataframe["Enc."] = self.real_dataframe["Enc."].replace(
                    0, "☐"
                )
                self.real_dataframe["Enc."] = self.real_dataframe["Enc."].replace(
                    1, "☒"
                )
                self.real_dataframe["Enc."] = self.real_dataframe["Enc."].replace(
                    "0", "☐"
                )
                self.real_dataframe["Enc."] = self.real_dataframe["Enc."].replace(
                    "1", "☒"
                )
                self.real_dataframe["Entr."] = self.real_dataframe["Entr."].replace(
                    0, "☐"
                )
                self.real_dataframe["Entr."] = self.real_dataframe["Entr."].replace(
                    1, "☒"
                )
                self.real_dataframe["Entr."] = self.real_dataframe["Entr."].replace(
                    "0", "☐"
                )
                self.real_dataframe["Entr."] = self.real_dataframe["Entr."].replace(
                    "1", "☒"
                )
                self.update_treeview()  # Update original tree
                add_window.destroy()  # close add_window
            else:
                messagebox.showerror("Erro", "Não existem elementos para adicionar!")

        def remove_addition():
            """
            Clears the added values from the treeview and from the added menu
            ´visual_edit´: Determines if the clear needs to remove actual files or if all
            it's doing is clearing the GUI of ´add_window´
            """
            visual_edit = 0
            try:
                df_added_old = pd.read_csv(path_add_csv)
            except:
                df_added_old = pd.DataFrame()
                visual_edit = 1  # If theres no csv to be deleted, then all we need to do is clear the xlsx
            list_desig = []

            if os.path.exists(path_add_csv) or os.path.exists(path_add_excel):
                confirmation_box = tk.Toplevel(root)
                confirmation_box.title("Alerta")
                confirmation_box.iconbitmap(icon)
                confirmation_box.resizable(0, 0)
                confirmation_box.attributes(
                    "-topmost", "true"
                ) 

                def yes_del():
                    try:
                        treeview_a.delete(*treeview_a.get_children()) 
                        if os.path.exists(path_add_excel):
                            os.remove(
                                path_add_excel
                            )  
                            new_addition(open_excel=0)  
                        if (
                            visual_edit == 0
                        ):  
                            if os.path.exists(path_add_csv):
                                os.remove(path_add_csv)  
                            for index, row in df_added_old.iterrows():
                                desig_row = row["Designação"]
                                list_desig.append(desig_row)
                                for (
                                    index_real,
                                    row_real,
                                ) in self.real_dataframe.iterrows():
                                    if row_real["Designação"] == desig_row:
                                        self.real_dataframe = self.real_dataframe.drop(
                                            index_real
                                        )
                            file_name = "controlo.csv"
                            con_path = os.path.join(csv_path, file_name)
                            if os.path.exists(con_path):
                                df_con = pd.read_csv(con_path)
                                for index, row in df_con.iterrows():
                                    if row["Designação"] in list_desig:
                                        df_con = df_con.drop(index)
                                df_con.to_csv(con_path, index=False)

                        confirmation_box.destroy()
                    except:
                        messagebox.showerror(
                            "Erro", "Feche o excel antes de limpar os dados!"
                        )

                def no_del():
                    confirmation_box.destroy()

                info_yesno = ttk.Label(
                    confirmation_box, text="Deseja remover os itens adicionados?"
                )
                info_yesno.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

                yes_button = ttk.Button(confirmation_box, text="Sim", command=yes_del)
                yes_button.grid(row=1, column=0, padx=5, pady=5)

                no_button = ttk.Button(confirmation_box, text="Não", command=no_del)
                no_button.grid(row=1, column=1, padx=5, pady=5)

                confirmation_box.wait_window() 
                self.update_treeview()
            else:
                messagebox.showerror(
                    "Erro", "Não existe nenhum ficheiro para eliminar!"
                )

        edit_button = ttk.Button(frame_buttons, text="Modificar", command=edit_addition)
        edit_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        import_button = ttk.Button(
            frame_buttons, text="Importar", command=import_addition
        )
        import_button.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        export_button = ttk.Button(
            frame_buttons, text="Exportar", command=export_addition
        )
        export_button.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        add_button = ttk.Button(frame_buttons, text="Guardar", command=add_addition)
        add_button.grid(row=0, column=4, padx=5, pady=5, sticky="ew")

        remove_button = ttk.Button(
            frame_buttons, text="Limpar", command=remove_addition
        )
        remove_button.grid(row=0, column=5, padx=5, pady=5, sticky="ew")


    def update_comp_button(self):
        """Updates the comparing filter variable and updates the tree"""
        self.compare.set(self.compare.get())
        self.update_treeview()

    def update_edit_button(self):
        """Updates the edit filter variable and updates the tree"""
        self.edited.set(self.edited.get())
        self.update_treeview()

    def update_remo_button(self):
        """Updates the removed filter variable and updates the tree"""
        self.removed.set(self.removed.get())
        self.update_treeview()

    def update_add_button(self):
        """Updates the added filter variable and updates the tree"""
        self.added.set(self.added.get())
        self.update_treeview()

    def on_combobox_change(self, event):
        """Updates the combobox variable"""
        self.dropdown_menu.selection_clear()
        self.update_treeview()

    def export_dataframe_to_excel(self):
        """Exports the dataframe in the GUI as a xlsx to be printed/analyzed by another party"""
        title = project_name

        excel_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )

        df = self.dataframe  

        with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:

            workbook = writer.book
            writer.sheets["Sheet1"] = workbook.create_sheet(index=0, title="Sheet1")
            worksheet = writer.sheets["Sheet1"]

            worksheet["A1"] = title
            worksheet.merge_cells("A1:H1")

            merged_cell = worksheet["A1"]
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")

            fill = PatternFill(
                start_color="92CDDC", end_color="92CDDC", fill_type="solid"
            )
            merged_cell.fill = fill

            bold_font = Font(bold=True)
            merged_cell.font = bold_font

            df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=1)
            header_style = NamedStyle(
                name="header_style",
                font=Font(bold=True, color="FFFFFF"),
                fill=PatternFill(
                    start_color="004080", end_color="004080", fill_type="solid"
                ),
                alignment=Alignment(horizontal="center", vertical="center"),
            )

            for cell in worksheet["2"]:
                cell.style = header_style

            for column in worksheet.columns:
                max_length = 0
                column_header = column[1].column_letter  
                column_data = [str(cell.value) for cell in column[1:]]
                for cell in [column[0]] + column_data: 
                    try:
                        if len(str(cell)) > max_length:
                            max_length = len(cell)
                    except:
                        pass

                adjusted_width = max_length + 5
                worksheet.column_dimensions[column_header].width = adjusted_width

            for row in worksheet.iter_rows():
                for cell in row:
                    border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                    cell.border = border

            if self.compare.get() == True:
                self.removed_items = self.rows_dif[2]
                self.added_items = self.rows_dif[1]
                self.edited_items = self.rows_dif[0]
                for row in worksheet.iter_rows():
                    fill_color = None
                    for cell in row:
                        cell_value = cell.value
                        if cell.value in self.removed_items:
                            fill_color = "ff7066"
                        elif cell.value in self.edited_items:
                            fill_color = "84ff4f"
                        elif cell.value in self.added_items:
                            fill_color = "ffb957"
                        if fill_color != None:
                            filter_fill = PatternFill(
                                start_color=fill_color,
                                end_color="bf1d00",
                                fill_type="solid",
                            )
                            for cell_in_row in row:
                                cell_in_row.fill = filter_fill
                                if cell_in_row.column == "G":
                                    break
            center_columns = ["C", "D", "E", "F", "G", "H"]
            for col_letter in center_columns:
                for cell in worksheet[col_letter]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    def update_treeview(self):
        """Updates the treeview depending on the options toggled on the GUI,
        every button press will trigger this function to run and update the GUI."""

        list_added_items = []
        if os.path.exists(add_path):
            df_add = pd.read_csv(add_path)
            for index, row in df_add.iterrows():
                list_added_items.append(row["Designação"])

        self.dataframe = (
            self.real_dataframe
        )  

        if self.compare.get() == False:
            self.removed_button["state"] = "disabled"
            self.edited_button["state"] = "disabled"
            self.added_button["state"] = "disabled"
            self.dropdown_menu["state"] = "disabled"
            self.dataframe = self.real_dataframe

        elif self.compare.get():
            actual_version = ((self.dropdown_var.get().split("-"))[0]).replace(" ", "")
            self.rows_dif = compare_data(self.real_dataframe, actual_version)
            self.removed_dataframe = self.rows_dif[3]


            self.dataframe = self.dataframe._append(self.removed_dataframe)
            self.dataframe = self.dataframe.fillna("")

            self.removed_button["state"] = "enabled"
            self.edited_button["state"] = "enabled"
            self.added_button["state"] = "enabled"
            self.dropdown_menu["state"] = "readonly"

            if self.compare.get() == True and (
                self.edited.get() == True
                or self.removed.get() == True
                or self.added.get() == True
            ):
                self.removed_items = self.rows_dif[2]
                self.added_items = self.rows_dif[1]
                self.edited_items = self.rows_dif[0]
                self.compare_dataframe = pd.DataFrame()

                if self.removed.get() == True:
                    for index, row in self.dataframe.iterrows():
                        if row["Designação"] in self.removed_items:
                            if row.to_dict() in self.compare_dataframe.to_dict(
                                orient="records"
                            ):
                                pass
                            else:
                                self.compare_dataframe = self.compare_dataframe._append(
                                    row
                                )

                if self.edited.get() == True:
                    for index, row in self.dataframe.iterrows():
                        if row["Designação"] in self.edited_items:
                            if row.to_dict() in self.compare_dataframe.to_dict(
                                orient="records"
                            ):
                                pass
                            else:
                                self.compare_dataframe = self.compare_dataframe._append(
                                    row
                                )
                if self.added.get() == True:
                    for index, row in self.dataframe.iterrows():
                        if row["Designação"] in self.added_items:
                            if row.to_dict() in self.compare_dataframe.to_dict(
                                orient="records"
                            ):
                                pass
                            else:
                                self.compare_dataframe = self.compare_dataframe._append(
                                    row
                                )

                self.dataframe = (
                    self.compare_dataframe
                )  # Make the GUI equal to the compare_dataframe

        if self.searching_bol == True:
            self.update_search()
            self.dataframe = self.filtered_dataframe
        elif self.searching_bol == False and self.compare.get() == False:
            self.dataframe = self.real_dataframe
        else:
            pass
        self.treeview.delete(*self.treeview.get_children())

        for index, row in self.dataframe.iterrows():
            values = list(row)
            item_id = self.treeview.insert("", "end", values=values)

            self.treeview.tag_configure("gray_row", background="#f0f0f0")
            self.treeview.item(item_id, tags="gray_row")
            if row["Designação"] in list_added_items:
                self.treeview.tag_configure(
                    "blue", foreground="#10397a", background="#f0f0f0"
                )
                self.treeview.item(item_id, tags="blue")
            if self.compare.get():
                self.removed_items = self.rows_dif[2]
                self.added_items = self.rows_dif[1]
                self.edited_items = self.rows_dif[0]
                if (
                    row["Designação"] in self.removed_items
                    and row["Designação"] not in list_added_items
                ):
                    self.treeview.tag_configure(
                        "red", foreground="#990000", background="#f0f0f0"
                    )
                    self.treeview.item(item_id, tags="red")
                elif (
                    row["Designação"] in self.added_items
                    and row["Designação"] not in list_added_items
                ):
                    self.treeview.tag_configure(
                        "green", foreground="#006600", background="#f0f0f0"
                    )
                    self.treeview.item(item_id, tag="green")
                elif (
                    row["Designação"] in self.edited_items
                    and row["Designação"] not in list_added_items
                ):
                    self.treeview.tag_configure(
                        "orange", foreground="#e08119", background="#f0f0f0"
                    )
                    self.treeview.item(item_id, tag="orange")
                else:
                    pass

    def edit_cell(self, event):
        """
        Function responsible for any permament changes in the dataframe
        Allows editing of certain fields
        """
        self.update_dialog = 0
        item = self.item_selected  
        self.desig_value = self.treeview.item(self.item_selected, "values")[1]
        for index, row in self.dataframe.iterrows():
            if self.desig_value in row["Designação"]:
                item_index = index
        column = self.treeview.identify_column(event.x)
        col_index = int(column.split("#")[-1]) - 1
        row = self.treeview.identify_row(event.x)
        value = self.treeview.item(item, "values")[col_index]
        enc_value = self.treeview.item(item, "values")[5]
        stock_value = self.treeview.item(item, "values")[4]
        code_value = self.treeview.item(item, "values")[0]
        desig_value = self.treeview.item(item, "values")[1]
        new_value = None
        date_value = self.treeview.item(item, "values")[6]
        add_path = os.path.join(main_path, "add.csv")
        list_added_items = []
        if os.path.exists(add_path):
            df_add = pd.read_csv(add_path)
            for index, row in df_add.iterrows():
                list_added_items.append(row["Designação"])

        if desig_value in list_added_items and col_index == 0:
            pass

        elif col_index != 7 and (value == "☐" or value == "☒"):
            if value == "☒":
                new_value = "☐"
            elif value == "☐":
                new_value = "☒"
        elif col_index == 1 or col_index == 2 or col_index == 3:
            pass

        elif enc_value == "☒" and col_index == 6:
            try:
                self.cal_dialog.destroy()
            except:
                pass
            self.cal_dialog = tk.Tk()
            cal = DateEntry(
                self.cal_dialog,
                date_pattern="dd-MM-yyyy",
                locale="pt",
                width=12,
                background="#4d4d4d",
                foreground="white",
                borderwidth=2,
            )
            if date_value != "" and date_value != None:
                cal.set_date(date_value)
            cal.pack(padx=0, pady=0)

            def set_date():
                """Simply used to extract the date from the calendar to the dataframe"""
                nonlocal new_value
                new_value = cal.get_date().strftime("%d-%m-%Y")
                self.cal_dialog.destroy()

            ok_button = Button(self.cal_dialog, text="   OK   ", command=set_date)

            ok_button.pack(pady=10)
            self.cal_dialog.update_idletasks()
            self.cal_dialog.wait_window(self.cal_dialog)
        elif enc_value == "☐" and col_index == 6:
            pass
        elif col_index == 7 and date_value != "" and enc_value == "☒":
            if value == "☒":
                new_value = "☐"
            elif value == "☐":
                new_value = "☒"

        elif col_index == 7 and enc_value == "☐":
            new_value = "☐"
        elif col_index == 7:
            pass
        else:

            def on_treeview_double_click(event):
                """Double click on a treeview item to insert the value into the simple input."""
                item = treeview.selection()[0]
                selected_value = treeview.item(item, "values")[
                    0
                ]
                simple_input_var.set(selected_value)
                add()

            def on_treeview_click(event):
                """Double click on a treeview item to insert the value into the simple input."""
                item = treeview.selection()[0]
                selected_value = treeview.item(item, "values")[
                    0
                ] 
                simple_input_var.set(selected_value)

            basic_edit_window = tk.Toplevel(self)
            basic_edit_window.title(f"Edição Código: {desig_value}")
            basic_edit_window.resizable(False, False)
            basic_edit_window.iconbitmap(icon)

            simple_input_label = tk.Label(basic_edit_window, text="Codigo:")
            simple_input_label.grid(row=2, column=0, sticky="e", padx=10, pady=5)

            simple_input_var = tk.StringVar(value=code_value)
            simple_input_entry = tk.Entry(
                basic_edit_window, textvariable=simple_input_var
            )
            simple_input_entry.grid(row=2, column=1, sticky="ew", padx=10, pady=5)

            def search(*args):
                """Search for matches in any column of the dataframe based on the input string."""
                search_string = search_input_var.get()
                filtered_df = self.df_phc[
                    self.df_phc.apply(
                        lambda row: any(
                            str(val).lower().count(search_string.lower()) for val in row
                        ),
                        axis=1,
                    )
                ]
                treeview.delete(*treeview.get_children())
                for index, row in filtered_df.iterrows():
                    values = [row[col] for col in tree_columns]
                    treeview.insert("", "end", values=values)

            search_input_label = tk.Label(basic_edit_window, text="Procurar:")
            search_input_label.grid(row=0, column=0, sticky="e", padx=10, pady=5)

            search_input_var = tk.StringVar(value=code_value)
            search_input_entry = tk.Entry(
                basic_edit_window, textvariable=search_input_var
            )
            search_input_entry.grid(
                row=0, column=1, columnspan=4, sticky="ew", padx=10, pady=5
            )
            search_input_var.trace_add("write", search)

            def add():
                self.update_dialog = 1
                self.new_value = simple_input_var.get()
                basic_edit_window.destroy()

            add_button = tk.Button(
                basic_edit_window, text="Adicionar Codigo", command=add
            )
            add_button.grid(row=2, column=2, sticky="ew", padx=10, pady=5)

            treeview_frame = tk.Frame(basic_edit_window)
            treeview_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=10)

            tree_columns = list(self.df_phc.columns)
            treeview_scrollbar_y = tk.Scrollbar(treeview_frame, orient="vertical")
            treeview = ttk.Treeview(
                treeview_frame,
                columns=tree_columns,
                show="headings",
                yscrollcommand=treeview_scrollbar_y.set,
            )

            for col in tree_columns:
                treeview.heading(col, text=col)

            for index, row in self.df_phc.iterrows():
                values = [row[col] for col in tree_columns]
                treeview.insert("", "end", values=values)

            for col in tree_columns:
                if col == "Ref":
                    treeview.column(col, width=80)
                if col == "Design":
                    treeview.column(col, width=500)

            treeview_scrollbar_y.config(command=treeview.yview)
            treeview_scrollbar_y.pack(side="right", fill="y")

            treeview.pack(fill="both", expand=True)

            treeview.bind("<Double-1>", on_treeview_double_click)
            treeview.bind("<ButtonRelease-1>", on_treeview_click)
            search()
            basic_edit_window.wait_window()

        if self.update_dialog == 1:
            new_value = self.new_value

        if new_value is not None:
            try:
                self.real_dataframe.iloc[item_index, col_index] = new_value
            except:
                pass
        if col_index == 5 and new_value == "☒" and stock_value == "☒":
            self.real_dataframe.iloc[item_index, 4] = "☐"
        elif col_index == 4 and new_value == "☒" and enc_value == "☒":
            self.real_dataframe.iloc[item_index, 5] = "☐"
            if date_value != None or date_value != "":
                self.real_dataframe.iloc[item_index, 6] = ""
        elif (
            col_index == 5
            and new_value == "☐"
            and (date_value != None or date_value != "")
        ):
            self.real_dataframe.iloc[item_index, 6] = ""
            self.real_dataframe.iloc[item_index, 7] = "☐"
        if self.searching_bol:
            self.update_search()
        self.update_treeview()

    def save_csv(self):
        """Save the csv/txt with the changes the user did"""
        file_name = "controlo.csv"
        file_path = os.path.join(csv_path, file_name)
        added_path = os.path.join(main_path, "add.csv")
        if os.path.exists(added_path):
            df_added_old = pd.read_csv(added_path)
        current_time = datetime.now()
        formatted_time = current_time.strftime("%d/%m/%Y %H:%M")
        if file_path:
            with open(log_path, "a") as file:
                file.write(f"\n_{last_v}:{username}, {formatted_time}")
            self.real_dataframe.replace("☐", "0", inplace=True)
            self.real_dataframe.replace("☒", "1", inplace=True)
            self.real_dataframe.to_csv(file_path, index=False)
            if os.path.exists(added_path):
                df_added = pd.merge(
                    self.real_dataframe,
                    df_added_old,
                    on="Designação",
                    how="inner",
                    suffixes=("", "_old"),
                )
                df_added.drop(
                    df_added.filter(like="_old").columns, axis=1, inplace=True
                )
                df_added.to_csv(added_path, index=False)
            df_clean = self.real_dataframe
            df_clean = df_clean.drop(
                ["Comp.(m)", "Qt.", "Stock", "Enc.", "Data Enc.", "Entr."], axis=1
            )
            self.df_comp = pd.concat([self.df_comp, df_clean])
            self.df_comp = self.df_comp.dropna(subset=["Codigo"])
            self.df_comp = self.df_comp[self.df_comp["Codigo"] != ""]
            self.df_comp = self.df_comp[self.df_comp["Codigo"].str.len() >= 8]
            self.df_comp = self.df_comp.drop_duplicates(
                keep="last", subset="Designação"
            )
            self.df_comp.rename(columns={"Designação": "Design"}, inplace=True)
            self.df_comp.to_excel(comp_excel, index=False)
            self.df_comp = pd.read_excel(comp_excel)
            self.df_comp.rename(columns={"Design": "Designação"}, inplace=True)
            tk.messagebox.showinfo("Sucesso", "Guardado!")

        txt_name = "notas.txt"
        txt_path = os.path.join(main_path, txt_name)
        text_content = self.note_text.get("1.0", tk.END)
        with open(txt_path, "w+") as file:
            file.write(text_content)

    def search_bol_update(self, *args):
        if self.search_var.get() is not None:
            query = self.search_var.get().lower()

        if query != "":
            self.searching_bol = True
        else:
            self.searching_bol = False
        self.update_treeview()

    def update_search(self, *args):
        """
        Toggles searching_bol true if the user types anything on the search bar
        The remaining logic is in self.update_treeview
        """

        if self.compare.get() == True and (
            self.edited.get() == True
            or self.removed.get() == True
            or self.added.get() == True
        ):
            self.dataframe = self.compare_dataframe
        else:
            self.dataframe = self.real_dataframe
        query = None
        try:
            if self.search_var.get() is not None:
                query = self.search_var.get().lower()
                if query != "" or query is not None:
                    self.filtered_dataframe = self.dataframe[
                        self.dataframe.apply(
                            lambda row: any(
                                str(val).lower().find(query) != -1 for val in row
                            ),
                            axis=1,
                        )
                    ]
        except:
            pass


if __name__ == "__main__":
    dir_path = os.path.dirname(os.path.realpath(__file__))
    root = DataFrameEditor(df_global)
    root.tk.call("source", os.path.join(dir_path, "azure.tcl"))
    root.tk.call("set_theme", "light")
    root.iconbitmap(icon)
    root.mainloop()
