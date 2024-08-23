import tkinter as tk
from tkinter import ttk, Button
from tkinter import scrolledtext, messagebox
from tkinter import filedialog
import os
from datetime import datetime, date
import sys
import subprocess
import shutil
import time
import threading
from tkcalendar import DateEntry
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side


script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
compras_py = os.path.join(script_dir, "compras", "compras.exe")
options = os.path.join(script_dir, "compras", "options.txt")
icon = os.path.join(script_dir, "compras", "md.ico")
with open(options, "r") as file:
    lines = file.readlines()
for line in lines:
    if "phc =" in line:
        folder_path_strip = line.split("= ", 1)[
            1
        ].strip()  
    if "user =" in line:
        username = line.split("=", 1)[1].strip()  
folder_path = os.path.join(folder_path_strip, "Obras")


def walklevel(some_dir, level):
    """Gives me subfolders x levels beyond the main folder,
    used to get the actual projects that are subfolders inside of multiple folders in the main path
    """
    some_dir = some_dir.rstrip(os.path.sep)
    assert os.path.isdir(some_dir)
    num_sep = some_dir.count(os.path.sep)
    for root, dirs, files in os.walk(some_dir):
        yield root, dirs, files
        num_sep_this = root.count(os.path.sep)
        if num_sep + level <= num_sep_this:
            del dirs[:]


def folder_to_data(path, archive=0):
    """Turns the list of folders into a dataframe to be turned into a treeview in the GUI"""
    df_folders = pd.DataFrame(
        columns=["Conjunto", "Utilizador", "Data Edição", "Verif."]
    )  
    folders = list(
        walklevel(path, level=2)
    )  
    skip = 0
    if len(folders) == 1:
        skip = 1
        df_enco = pd.DataFrame()

    if skip == 0:

        list_folder = []
        paths = []
        paths_list = []
        max_path_len = 0
        project_paths = []
        for folder in folders:
            list_folder.append(folder[0])
            for (
                pathies
            ) in (
                list_folder
            ):  
                paths = pathies.split("\\")
                if len(paths) > max_path_len:
                    max_path_len = len(
                        paths
                    )  
                paths_list.append(paths)  
        project_paths = [
            x for x in paths_list if len(x) >= max_path_len
        ]  
        project_paths = [
            i for n, i in enumerate(project_paths) if i not in project_paths[:n]
        ] 
        project_paths_real = []  
        for (
            paths
        ) in project_paths:  
            paths[0] = paths[0] + "\\"
            path = os.path.join(*paths)
            project_paths_real.append(path)
        for (
            path
        ) in (
            project_paths_real
        ):  
            try:
                log_path = os.path.join(path, "log.txt")
                if os.path.exists(log_path):
                    with open(log_path, "r") as f:
                        lines = f.readlines()
                name_project = (
                    path.split("\\")[-2] + ": " + path.split("\\")[-1]
                )  
                verified = None
                if lines[-1].startswith("V"):
                    verified = 0
                elif lines[-1].startswith("."):
                    verified = 2
                else:
                    verified = 1

                last_user = (
                    (lines[-1].split(":"))[1].split(",")[0]
                ).split()  

                last_date = ((lines[-1].split(":")[1]).split(",")[1])[
                    1:11
                ] 

                list_df = [
                    name_project,
                    last_user,
                    last_date,
                    verified,
                ]  

                if archive == 1 or (archive == 0 and verified != 2):
                    df_folders = df_folders._append(
                        pd.Series(list_df, index=df_folders.columns), ignore_index=True
                    ) 

            except:  
                root = tk.Tk()
                root.withdraw()  
                messagebox.showerror(
                    "Erro",
                    "Elimine a pasta e contacte o Projeto para reenviar!\nPasta corrompida: {}".format(
                        path
                    ),
                )

        df_folders["Data Edição"] = pd.to_datetime(
            df_folders["Data Edição"], format="%d/%m/%Y"
        ).dt.date

       
        df_folders = df_folders.sort_values(by="Data Edição", ascending=False)
        df_enco = pd.DataFrame()
        last_review_bool = 0  

        for index, row in df_folders.iterrows():

            name = row["Conjunto"]
            name_1 = name.split(":")[0]
            name_2 = name.split(": ")[1]
            log_path = os.path.join(folder_path, name_1, name_2, "log.txt")
            with open(log_path, "r") as f:
                lines = f.readlines()
            last_review_v = None
            for (
                line
            ) in (
                lines
            ):  
                if line.startswith("_"):
                    last_review_v = line.split(":")[0][1:]
            if (
                last_review_v != None
            ):  
                last_review_bool = (
                    1  # Set it to 1 to represent that theres actual information
                )
                control_csv_path = os.path.join(
                    folder_path, name_1, name_2, last_review_v, "CSV", "controlo.csv"
                )
                df_control = pd.read_csv(control_csv_path)
                df_control.dropna(
                    subset=["Data Enc."], inplace=True
                )  
                df_control = df_control.drop(
                    ["Stock", "Enc.", "Comp.(m)"], axis=1
                )  
                df_control["Obra"] = name
                df_enco = pd.concat(
                    [df_enco, df_control], ignore_index=True
                )  

        if last_review_bool == 1:
            df_enco["Data Enc."] = pd.to_datetime(
                df_enco["Data Enc."], format="%d-%m-%Y"
            )
            df_enco = df_enco.sort_values(by="Data Enc.")
            df_enco["Data Enc."] = df_enco["Data Enc."].dt.strftime("%d-%m-%Y")
            df_enco.reset_index(drop=True, inplace=True)
            df_enco["Data Enc."] = df_enco["Data Enc."].astype(str)
            df_enco["Entr."] = df_enco["Entr."].astype(str)
            df_enco["Entr."] = df_enco["Entr."].replace("0", "☐")
            df_enco["Entr."] = df_enco["Entr."].replace("1", "☒")
            df_enco.fillna("", inplace=True)
        df_folders.loc[:, "Verif."] = df_folders["Verif."].replace(2, "X")
        df_folders.loc[:, "Verif."] = df_folders["Verif."].replace(1, "✓")
        df_folders.loc[:, "Verif."] = df_folders["Verif."].replace(0, "")

    return df_folders, df_enco


class FolderTreeViewApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Compras/Produção")
        self.root.iconbitmap(icon)
        self.root.resizable(False, False)  
        self.archive_value = 0  
        self.df_tree = folder_to_data(folder_path, self.archive_value)
        self.df_folders = self.df_tree[0]
        self.df_enco = self.df_tree[1]
        self.export_list = []
        self.frame1 = ttk.Frame(root, padding="10")
        self.frame1.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.list_proj = [""]
        self.archive_filt = tk.BooleanVar(value=False)

        scrollbar = ttk.Scrollbar(self.frame1, orient="vertical")
        scrollbar.grid(row=0, column=1, sticky="nsew")

        self.treeview = ttk.Treeview(
            self.frame1,
            columns=list(self.df_folders.columns),
            show="headings",
            yscrollcommand=scrollbar.set,
        )
        
        scrollbar.config(command=self.treeview.yview)

        self.treeview.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        for col in self.df_folders.columns:
            self.treeview.heading(col, text=col)


        for col in self.df_folders.columns:
            try:
                max_width = (
                    max(self.df_folders[col].astype(str).apply(lambda x: len(x))) + 4
                )
            except:
                max_width = 20
            padding = 4  
            self.treeview.column(col, width=(max_width + padding) * 7)
            if col == "Data Edição":
                self.treeview.column(col, width=80)
            if col == "Verif.":
                self.treeview.column(col, width=45)

        all_columns = []
        for col in self.df_folders.columns:
            if col == "Verif." or col == "Data Edição" or col == "Utilizador":
                self.treeview.heading(col, text=col, anchor=tk.CENTER)
                self.treeview.column(col, anchor=tk.CENTER)
            all_columns.append(col)
        self.treeview["height"] = 20
        for index, row in self.df_folders.iterrows():
            self.treeview.insert("", "end", text=index, values=list(row))

        self.treeview.bind("<ButtonRelease-1>", self.on_treeview_click)
        self.treeview.bind("<Double-Button-1>", self.on_treeview_d_click)

        self.frame2 = ttk.Frame(root, padding="10")
        self.frame2.grid(row=0, column=1, rowspan=2, sticky="nsew")

        def search(*args):
            """Search for matches in any column of the dataframe based on the input string."""
            search_string = search_input_var.get()
            filtered_df = self.df_folders[
                self.df_folders.apply(
                    lambda row: any(
                        str(val).lower().count(search_string.lower()) for val in row
                    ),
                    axis=1,
                )
            ]
            self.treeview.delete(*self.treeview.get_children())
            for index, row in filtered_df.iterrows():
                values = [row[col] for col in all_columns]
                self.treeview.insert("", "end", values=values)

        larger_font = (
            "Arial",
            12,
        )  

        search_input_label = tk.Label(self.frame2, text="Procurar:", font=larger_font)
        search_input_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)
        search_input_var = tk.StringVar(value="")
        search_input_entry = tk.Entry(self.frame2, textvariable=search_input_var)
        search_input_entry.grid(
            row=0, column=1, columnspan=4, sticky="ew", padx=5, pady=5
        )
        search_input_var.trace_add(
            "write", search
        )  

        self.print_box = scrolledtext.ScrolledText(
            self.frame2, wrap=tk.WORD, width=55, height=20
        )
        self.print_box.grid(row=3, column=0, columnspan=3)
        self.print_box["state"] = (
            "disabled"  
        )

        self.export_button = ttk.Button(
            self.frame2, text="Exportar", command=self.export
        )
        self.export_button.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.archive_button = ttk.Button(
            self.frame2, text="Arquivar", command=self.archive
        )
        self.archive_button.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        self.archive_filter = ttk.Checkbutton(
            self.frame2,
            text="Arquivos",
            variable=self.archive_filt,
            command=self.archive_switch,
        )
        self.archive_filter.grid(row=2, column=1, padx=5, pady=5)

        self.calendar_button = ttk.Button(
            self.frame2, text="Calendario", command=self.calendar
        )
        self.calendar_button.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        self.setting_button = ttk.Button(
            self.frame2, text="Definições", command=self.settings
        )
        self.setting_button.grid(row=1, column=2, padx=5, pady=5, sticky="ew")

        self.root.rowconfigure(1, weight=0, minsize=3)

        self.up_t = threading.Thread(target=self.update)
        self.up_t.daemon = True
        self.up_t.start()

    def archive(self):
        """
        Archives projects so that they are not accounted by the calendar and not visible by the user
        It simply adds another line to the log.txt that makes them be considered archived
        if the last line of the log.txt starts with '.' the project will be considered archived
        saving it/sending a new version will remove that state as it adds another line to the log.txt
        """
        selected_item = (
            self.treeview.selection()
        ) 
        if selected_item:
            item_values = self.treeview.item(selected_item, "values")
            first_folder = (item_values[0].split(": "))[0]
            second_folder = (item_values[0].split(": "))[1]
            log_path = os.path.join(folder_path, first_folder, second_folder, "log.txt")
            current_time = datetime.now()
            formatted_time = current_time.strftime("%d/%m/%Y %H:%M")
            with open(log_path, "r") as file:
                lines = [line.strip() for line in file]
            last_version = None
            for line in lines:
                if line[0] == "V":  
                    last_version = line
                else:
                    pass
                last_v = last_version.split(":")[0]
            with open(
                log_path, "a"
            ) as file:  
                file.write(f"\n.{last_v}:{username}, {formatted_time}")

    def archive_switch(self):
        """Switch so the program takes into consideration again archived projects"""
        self.archive_filt.set(self.archive_filt.get())
        if self.archive_value == 0:
            self.archive_value = 1
        else:
            self.archive_value = 0

    def update(self):
        """Updates the treeview if there is a change in data"""
        while True:
            new_df_tree = folder_to_data(folder_path, self.archive_value)
            new_df_folders = new_df_tree[0]
            if not new_df_folders.equals(self.df_folders):
                self.df_tree = new_df_tree
                self.df_folders = self.df_tree[0]
                self.df_enco = self.df_tree[1]
  
                for item in self.treeview.get_children():
                    self.treeview.delete(item)
     
                for index, row in self.df_folders.iterrows():
                    values = list(row)
                    self.treeview.insert("", "end", text=index, values=values)
                print("Treeview updated")
            time.sleep(1)

    def settings(self):
        """Opens a window that changes the values in the options.txt"""
        setting_window = tk.Toplevel(self.root)
        setting_window.title("Definições")
        setting_window.iconbitmap(icon)
        setting_window.resizable(False, False)

        frame1 = ttk.Frame(setting_window)
        frame1.grid(row=0, column=0, padx=0, pady=10, sticky="nsew")

        with open(options, "r") as file:
            lines = file.readlines()
            for line in lines:
                if line.startswith("user ="):
                    user_line = line
                if line.startswith("time ="):
                    time = (line.split("= ", 1)[1]).strip()
        user_name = user_line.split("= ")[1]

        notification_time_minutes = int(int(time) / 60)

        def path_data():
            """Open a file dialog for the user to choose the path to the folder"""
            path_choosen_by_user = filedialog.askdirectory(title="Selecione o caminho")
            if path_choosen_by_user:
                phc_path = os.path.join(path_choosen_by_user, "phc.xlsx")
                if os.path.exists(phc_path):
     
                    with open(options, "r") as file:
                        lines = file.readlines()
          
                    if lines:
                        lines[2] = f"phc = {path_choosen_by_user}\n"
     
                    with open(options, "w") as file:
                        file.writelines(lines)
                else:
                    messagebox.showerror(
                        "Erro",
                        'Selecione a pasta correta "Dados PHC", \nA pasta encontra-se no servidor na pasta de produção!',
                    )
            else:
                messagebox.showerror("Erro", "Selecione uma pasta")

        path_button = ttk.Button(
            setting_window, text="Definir Caminho Produção", command=path_data
        )
        path_button.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

        def user_update(*args):
            """updates log.txt lines"""
            user_real_value = user_input_var.get()
            try:
                noti_real = int(noti_input_var.get())
                noti_real = noti_real * 60
            except:
                noti_real = 3600
            with open(options, "r") as file:
                lines = file.readlines()
            updated_lines = []
            for line in lines:
                if line.startswith("user ="):
                    user_value = line.split("= ")[1]
                    print(
                        f"user_value: {user_value} | user_real_value: {user_real_value}"
                    )
                    if user_value == user_real_value:
                        updated_lines.append(f"user = {user_real_value}")
                    else:
                        updated_lines.append(f"user = {user_real_value}\n")
                elif line.startswith("time ="):
                    updated_lines.append(f"time = {noti_real}\n")
                else:
                    updated_lines.append(line)
            with open(options, "w") as file:
                file.writelines(updated_lines)

        def apro_open():
            aprov_path = os.path.join(folder_path_strip, "aproveitamento.xlsx")
            os.startfile(aprov_path)

        def dxf_open():
            dxf_path = os.path.join(folder_path_strip, "dxf_ana.xlsx")
            os.startfile(dxf_path)

        user_input_label = tk.Label(setting_window, text="Utilizador:")
        user_input_label.grid(row=0, column=0, sticky="e", padx=10, pady=5)
        noti_input_label = tk.Label(
            setting_window, text="Notificações Temporizador (minutos):"
        )
        noti_input_label.grid(row=1, column=0, columnspan=2, padx=10, pady=5)

        user_input_var = tk.StringVar(value=user_name)
        user_input_entry = tk.Entry(
            setting_window, textvariable=user_input_var, width=25
        )

        noti_input_var = tk.StringVar(value=notification_time_minutes)
        noti_input_entry = tk.Entry(
            setting_window, textvariable=noti_input_var, width=25
        )

        user_input_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=5)
        noti_input_entry.grid(
            row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=5
        )

        apro_button = ttk.Button(
            setting_window, text="Aproveitamento Perfis", command=apro_open
        )
        apro_button.grid(row=3, column=0, columnspan=2, sticky="ew", padx=10, pady=5)

        dxf_button = ttk.Button(
            setting_window, text="Regras de Corte DXF", command=dxf_open
        )
        dxf_button.grid(row=4, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        user_input_var.trace_add("write", user_update)
        noti_input_var.trace_add("write", user_update)

    def calendar(self):
        """Creates a calendar(treeview) GUI that shows all deliveries that were delivered or to be delivered,
        allows the user to cancel deliveries, change date of delivery and change the delivery state
        allows filtering of things yet to be delivered and specific projects"""
        
        if len(self.df_enco) >= 1:
            basic_edit_window = tk.Toplevel(self.root)
            basic_edit_window.title(f"Calendario")
            basic_edit_window.iconbitmap(icon)
            basic_edit_window.resizable(
                False, False
            )  

            frame1 = ttk.Frame(basic_edit_window)
            frame1.grid(row=0, column=0, padx=0, pady=10, sticky="nsew")

            self.df_calendar = self.df_enco[self.df_enco["Entr."] != "☒"]
            self.df_calendar_state = 0

            def on_treeview_singleclick(event):
                """updates the selected_item variable, mainly used for the remove function"""
                try:
                    selected_item = treeview.selection()[0]
                except:
                    pass

            def on_treeview_doubleclick(event):
                """Allows change of dates of delivery and state of delivery by double clicking"""

                selected_item = treeview.selection()[0] 
                column = treeview.identify_column(event.x)  
                col_index = int(column.split("#")[-1]) - 1  
                value = treeview.item(selected_item, "values")[
                    col_index
                ] 
                desig_tree = treeview.item(selected_item, "values")[
                    1
                ]  
                date_tree = treeview.item(selected_item, "values")[3]
                
                obra_tree = treeview.item(selected_item, "values")[5]
                path_1 = obra_tree.split(":")[0]
                path_2 = obra_tree.split(": ")[1]
                log_path = os.path.join(folder_path, path_1, path_2, "log.txt")

                
                last_review = None

                with open(log_path, "r") as file:
                    lines = file.readlines()
                for line in lines:
                    if line.startswith("_"):
                        last_review = line.split(":")[0][1:]

                
                controlo_path = os.path.join(
                    folder_path, path_1, path_2, last_review, "CSV", "controlo.csv"
                )
                path_add_csv = os.path.join(folder_path, path_1, path_2, "add.csv")
                df_controlo = pd.read_csv(controlo_path)
                df_controlo_unchanged = df_controlo
                new_value = None
                if os.path.exists(path_add_csv):
                    df_add = pd.read_csv(path_add_csv)

                
                if col_index == 4 or col_index == 3:
                    if col_index == 4:  
                        if value == "☒":
                            new_value = 0
                        elif value == "☐":
                            new_value = 1
                        for index, row in df_controlo.iterrows():
                            if desig_tree == row["Designação"]:
                                if new_value == None:
                                    new_value = df_controlo.at[index, "Entr."]
                                df_controlo.at[index, "Entr."] = new_value

                    if col_index == 3:

                        self.cal_dialog = tk.Tk()
                        cal = DateEntry(
                            self.cal_dialog,
                            date_pattern="dd-MM-yyyy",
                            locale="pt",
                            width=12,
                            background="#4d4d4d",
                            foreground="white",
                            borderwidth=2,
                        )
                        cal.set_date(date_tree)
                        cal.pack(padx=0, pady=0)

                        def set_date():
                            """Simply used to extract the date from the calendar to the dataframe"""
                            nonlocal new_value
                            new_value = cal.get_date().strftime("%d-%m-%Y")
                            self.cal_dialog.destroy()

                        ok_button = Button(
                            self.cal_dialog, text="   OK   ", command=set_date
                        )
                        ok_button.pack(pady=10)
                        self.cal_dialog.update_idletasks()
                        self.cal_dialog.wait_window(self.cal_dialog)
                        for index, row in df_controlo.iterrows():
                            if desig_tree == row["Designação"]:
                                if new_value == None:
                                    new_value = df_controlo.at[index, "Data Enc."]
                                df_controlo.at[index, "Data Enc."] = new_value
                    df_controlo.to_csv(
                        controlo_path, index=False
                    )  
                    if os.path.exists(path_add_csv):
                        for index, row in df_controlo.iterrows():
                            for index_a, row_a in df_add.iterrows():
                                if row_a["Designação"] == row["Designação"]:
                                    df_add.loc[index_a] = row
                        df_add.to_csv(path_add_csv, index=False)
                    
                    self.df_tree = folder_to_data(folder_path, self.archive_value)
                    self.df_enco = self.df_tree[1]

                    if self.df_calendar_state == 0:
                        self.df_calendar = self.df_enco[self.df_enco["Entr."] != "☒"]
                    elif self.df_calendar_state == 1:
                        self.df_calendar = self.df_enco

                    treeview.delete(
                        *treeview.get_children()
                    ) 
                    for index, row in self.df_calendar.iterrows():
                        values = list(row)
                        treeview.insert("", "end", text=index, values=values)

                    search()

            def search(*args):
                """Search for matches in any column of the dataframe based on the input string.
                also used as an updater"""

                self.list_proj = [""]

                
                for index, row in self.df_calendar.iterrows():
                    if row["Obra"] not in self.list_proj:
                        project = row["Obra"]
                        self.list_proj.append(project)
                    else:
                        pass

                proj_drop["values"] = self.list_proj  
                proj_drop.selection_clear()  
                proj_val = proj_var.get()

                search_string = search_input_var.get()
                filtered_df = self.df_calendar[
                    self.df_calendar.apply(
                        lambda row: any(
                            str(val).lower().count(search_string.lower()) for val in row
                        ),
                        axis=1,
                    )
                ]
                if proj_val != "" or None:
                    filtered_df = filtered_df[
                        filtered_df.apply(
                            lambda row: any(
                                str(val).lower().count(proj_val.lower()) for val in row
                            ),
                            axis=1,
                        )
                    ]
                
                treeview.tag_configure(
                    "orange", foreground="#e08119", background="#f0f0f0"
                )
                treeview.tag_configure(
                    "green", foreground="#006600", background="#f0f0f0"
                )
                treeview.delete(*treeview.get_children())
                for index, row in filtered_df.iterrows():
                    values = [row[col] for col in tree_columns]
                    item_id = treeview.insert("", "end", values=values)

                    today = date.today()
                    date_enc_str = row["Data Enc."]

                    if date_enc_str:
                        date_enc = datetime.strptime(date_enc_str, "%d-%m-%Y").date()
                        if row["Entr."] == "☒":
                            treeview.item(item_id, tag="green")
                        elif date_enc < today:
                            treeview.item(item_id, tag="orange")

            def entregues():
                """update entregues dropbox"""
                if self.df_calendar_state == 0:
                    self.df_calendar = self.df_enco
                    self.df_calendar_state = 1
                elif self.df_calendar_state == 1:
                    self.df_calendar = self.df_enco[self.df_enco["Entr."] != "☒"]
                    self.df_calendar_state = 0

                treeview.delete(*treeview.get_children())

                
                for index, row in self.df_calendar.iterrows():
                    values = list(row)
                    treeview.insert("", "end", text=index, values=values)
                search()

            def remove_enc():
                """Function to remove a delivery completly,
                this will clean all delivery related columns"""
                selected_item = treeview.selection()[0]  
                desig_tree = treeview.item(selected_item, "values")[
                    1
                ] 
                obra_tree = treeview.item(selected_item, "values")[5]
                path_1 = obra_tree.split(":")[0]
                path_2 = obra_tree.split(": ")[1]
                log_path = os.path.join(folder_path, path_1, path_2, "log.txt")

                
                last_review = None

                with open(log_path, "r") as file:
                    lines = file.readlines()
                for line in lines:
                    if line.startswith("_"):
                        last_review = line.split(":")[0][1:]

                
                controlo_path = os.path.join(
                    folder_path, path_1, path_2, last_review, "CSV", "controlo.csv"
                )

                df_controlo = pd.read_csv(controlo_path)  
                date = ""
                enc_val = "0"
                for index, row in df_controlo.iterrows():
                    if desig_tree == row["Designação"]:
                        df_controlo.at[index, "Data Enc."] = date
                        df_controlo.at[index, "Enc."] = enc_val

                df_controlo.to_csv(
                    controlo_path, index=False
                ) 
                self.df_tree = folder_to_data(folder_path, self.archive_value)
                self.df_enco = self.df_tree[1]

                if self.df_calendar_state == 0:
                    self.df_calendar = self.df_enco[self.df_enco["Entr."] != "☒"]
                elif self.df_calendar_state == 1:
                    self.df_calendar = self.df_enco

                treeview.delete(
                    *treeview.get_children()
                )  
                for index, row in self.df_calendar.iterrows():
                    values = list(row)
                    treeview.insert("", "end", text=index, values=values)

                search() 

            
            search_input_label = tk.Label(frame1, text="Procurar:")
            search_input_label.grid(row=0, column=0, sticky="e", padx=10, pady=5)

            search_input_var = tk.StringVar()
            search_input_entry = tk.Entry(
                frame1, textvariable=search_input_var, width=30
            )
            search_input_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=5)
            search_input_var.trace_add("write", search)

            
            delivered_button = ttk.Checkbutton(
                frame1, text=" Entregues ", command=entregues
            )
            delivered_button.grid(row=0, column=3, padx=10, pady=5, sticky="e")

            proj_var = tk.StringVar()
            proj_drop = ttk.Combobox(
                frame1,
                values=self.list_proj,
                state="readonly",
                textvariable=proj_var,
                width=30,
            )
            proj_drop.grid(row=0, column=4, padx=5, pady=5)
            proj_drop.bind("<<ComboboxSelected>>", search)

            remove_button = ttk.Button(
                frame1, text="Remover Encomenda", command=remove_enc
            )
            remove_button.grid(row=0, column=5, padx=10, pady=0, sticky="e")

            
            treeview_frame = tk.Frame(basic_edit_window)
            treeview_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=10)

            tree_columns = list(self.df_enco.columns)
            treeview_scrollbar_y = tk.Scrollbar(treeview_frame, orient="vertical")
            treeview = ttk.Treeview(
                treeview_frame,
                columns=tree_columns,
                show="headings",
                yscrollcommand=treeview_scrollbar_y.set,
            )

            
            for col in tree_columns:
                max_width = (
                    max(self.df_enco[col].astype(str).apply(lambda x: len(x))) + 4
                )
                padding = 4  
                treeview.column(col, width=(max_width + padding) * 7)
                treeview.heading(col, text=col)
                treeview.column(col, width=(max_width + padding) * 7)
                if col == "Codigo":
                    treeview.column(col, width=100)
                if col == "Ref":
                    treeview.column(col, width=100)
                if col == "Design":
                    treeview.column(col, width=500)
                elif col == "Entr.":
                    treeview.column(col, width=45)
                elif col == "Qt.":
                    treeview.column(col, width=60)
                elif col == "Data Enc.":
                    treeview.column(col, width=90)
                if (
                    col == "Qt."
                    or col == "Data Enc."
                    or col == "Entr."
                    or col == "Obra"
                ):
                    treeview.heading(col, text=col, anchor=tk.CENTER)
                    treeview.column(col, anchor=tk.CENTER)
            for index, row in self.df_enco.iterrows():
                values = [
                    (
                        row[col].strftime("%Y-%m-%d")
                        if isinstance(row[col], datetime)
                        else row[col]
                    )
                    for col in tree_columns
                ]
                treeview.insert("", "end", values=values)

            
            treeview_scrollbar_y.config(command=treeview.yview)
            treeview_scrollbar_y.pack(side="right", fill="y")

            treeview.pack(fill="both", expand=True)

            treeview.bind("<ButtonRelease-1>", on_treeview_singleclick)
            treeview.bind("<Double-1>", on_treeview_doubleclick)

            
            search()
            basic_edit_window.wait_window()
        else:
            messagebox.showerror("Erro", "Não existem encomendas")

    def export(self):
        """
        Opens a window that allows the user to export csv/excel/project files of projects
        """
        selected_item = self.treeview.selection()
        self.export_list = []
        if (
            selected_item
        ):  
            item_values = self.treeview.item(selected_item, "values")

            first_folder = (item_values[0].split(": "))[0]
            second_folder = (item_values[0].split(": "))[1]
            log_path = os.path.join(folder_path, first_folder, second_folder, "log.txt")

            with open(log_path, "r") as file:
                lines = file.readlines()
            dropdown_list = []

            for line in lines:
                if line.startswith("V"):
                    dropdown_value = line.split(":")[0]
                    dropdown_list.append(dropdown_value)

            dropdown_list = list(dict.fromkeys(dropdown_list))

            def update_exports(selected_val):
                """Update combobox values of the export window"""
                self.export_list = []
                version_path = os.path.join(
                    folder_path, first_folder, second_folder, selected_val
                )
                project_path = os.path.join(version_path, "Projeto")
                controlo_path = os.path.join(version_path, "CSV", "controlo.csv")
                if os.path.exists(project_path):
                    self.export_list.append("Dados Projeto")
                if os.path.exists(controlo_path):
                    self.export_list.append("Excel")
                    self.export_list.append("CSV")

                def on_combobox_change_2(event):
                    """Clears combobox selection higlight"""
                    dropdown_menu_2.selection_clear()

                try:
                    project_val = self.export_list[0]
                except:
                    project_val = ""
                self.export_method = tk.StringVar(value=project_val)
                dropdown_menu_2 = ttk.Combobox(
                    export_window,
                    values=self.export_list,
                    state="readonly",
                    textvariable=self.export_method,
                )
                dropdown_menu_2.grid(row=1, column=1, padx=5, pady=5)
                dropdown_menu_2.bind("<<ComboboxSelected>>", on_combobox_change_2)

            def on_combobox_change(event):
                selected_value = dropdown_var.get()
                dropdown_menu.selection_clear()
                update_exports(selected_value)

            export_window = tk.Toplevel(self.root)
            export_window.title(f"{first_folder}-{second_folder}")
            export_window.iconbitmap(icon)
            export_window.resizable(False, False)

            ver_label = ttk.Label(export_window, text="Versão")
            ver_label.grid(row=0, column=0, padx=5, pady=5)

            type_label = ttk.Label(export_window, text="Ficheiro")
            type_label.grid(row=0, column=1, padx=5, pady=5)

            dropdown_var = tk.StringVar(value=dropdown_value)
            dropdown_menu = ttk.Combobox(
                export_window,
                values=dropdown_list,
                state="readonly",
                textvariable=dropdown_var,
            )
            dropdown_menu.grid(row=1, column=0, padx=5, pady=5)
            dropdown_menu.bind("<<ComboboxSelected>>", on_combobox_change)

            selected_value = dropdown_var.get()
            update_exports(selected_value)

            def export_documents():
                """Exports different files depending on the value chosen by the user"""
                export_m = self.export_method.get()
                selected_v = dropdown_var.get()
                version_path = os.path.join(
                    folder_path, first_folder, second_folder, selected_v
                )
                project_path = os.path.join(version_path, "Projeto")
                controlo_path = os.path.join(version_path, "CSV", "controlo.csv")

                if export_m == "Dados Projeto":
                    
                    export_folder = filedialog.askdirectory(
                        title="Selecione onde exportar os ficheiros"
                    )
                    if export_folder:
                        
                        destination_folder_name = (
                            f"{first_folder}_{second_folder}_{selected_v}"
                        )
                        destination_path = os.path.join(
                            export_folder, destination_folder_name
                        )
                        try:
                            
                            shutil.copytree(project_path, destination_path)
                        except Exception as e:
                            pass
                if export_m == "CSV":
                    
                    export_folder = filedialog.askdirectory(
                        title="Selecione onde exportar o CSV"
                    )
                    if export_folder:
                        
                        destination_folder_name = (
                            f"{first_folder}_{second_folder}_{selected_v}.csv"
                        )
                        destination_path = os.path.join(
                            export_folder, destination_folder_name
                        )
                        try:
                            
                            shutil.copy(controlo_path, destination_path)
                        except Exception as e:
                            pass

                if export_m == "Excel":
                    export_folder = filedialog.askdirectory(
                        title="Selecione onde exportar o excel"
                    )

                    if export_folder:
                        destination_folder_name = (
                            f"{first_folder}_{second_folder}_{selected_v}.xlsx"
                        )
                        destination_path = os.path.join(
                            export_folder, destination_folder_name
                        )
                        df = pd.read_csv(controlo_path)

                        df["Stock"] = df["Stock"].astype(str)
                        df["Stock"] = df["Stock"].replace("0", "☐")
                        df["Stock"] = df["Stock"].replace("1", "☒")

                        df["Enc."] = df["Enc."].astype(str)
                        df["Enc."] = df["Enc."].replace("0", "☐")
                        df["Enc."] = df["Enc."].replace("1", "☒")

                        df["Entr."] = df["Entr."].astype(str)
                        df["Entr."] = df["Entr."].replace("0", "☐")
                        df["Entr."] = df["Entr."].replace("1", "☒")

                        with pd.ExcelWriter(
                            destination_path, engine="openpyxl"
                        ) as writer:
                            workbook = writer.book
                            writer.sheets["Sheet1"] = workbook.create_sheet(
                                index=0, title="Sheet1"
                            )
                            worksheet = writer.sheets["Sheet1"]

                            worksheet["A1"] = (
                                f"{first_folder}:{second_folder}({selected_v})"
                            )
                            worksheet.merge_cells("A1:H1")

                            merged_cell = worksheet["A1"]
                            merged_cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )

                            fill = PatternFill(
                                start_color="92CDDC",
                                end_color="92CDDC",
                                fill_type="solid",
                            )
                            merged_cell.fill = fill

                            
                            bold_font = Font(bold=True)
                            merged_cell.font = bold_font

                            
                            df.to_excel(
                                writer, sheet_name="Sheet1", index=False, startrow=1
                            )
                            header_style = NamedStyle(
                                name="header_style",
                                font=Font(bold=True, color="FFFFFF"),
                                fill=PatternFill(
                                    start_color="004080",
                                    end_color="004080",
                                    fill_type="solid",
                                ),  
                                alignment=Alignment(
                                    horizontal="center", vertical="center"
                                ),
                            )

                            
                            for cell in worksheet["2"]:
                                cell.style = header_style

                            
                            for column in worksheet.columns:
                                max_length = 0
                                column_header = column[1].column_letter  
                                column_data = [str(cell.value) for cell in column[1:]]
                                for cell in [
                                    column[0]
                                ] + column_data:
                                    try:
                                        if len(str(cell)) > max_length:
                                            max_length = len(cell)
                                    except:
                                        pass

                                adjusted_width = max_length + 5
                                worksheet.column_dimensions[column_header].width = (
                                    adjusted_width
                                )

                            
                            for row in worksheet.iter_rows():
                                for cell in row:
                                    border = Border(
                                        left=Side(style="thin"),
                                        right=Side(style="thin"),
                                        top=Side(style="thin"),
                                        bottom=Side(style="thin"),
                                    )
                                    cell.border = border
                            
                            center_columns = ["C", "D", "E", "F", "G", "H"]
                            for col_letter in center_columns:
                                for cell in worksheet[col_letter]:
                                    cell.alignment = Alignment(
                                        horizontal="center", vertical="center"
                                    )
                export_window.destroy()

            export_button = ttk.Button(
                export_window, text="Exportar", command=export_documents
            )
            export_button.grid(
                row=2, column=0, columnspan=2, padx=5, pady=5, sticky="ew"
            )

        else:
            messagebox.showerror("Erro", "Selecione um projeto")

    def run(self):
        self.root.mainloop()

    def on_treeview_click(self, event):
        """
        Updates the output window with the history of changes in the clicked project
        """
        selected_item = self.treeview.selection()
        if selected_item:
            item_values = self.treeview.item(selected_item, "values")

            
            first_folder = (item_values[0].split(": "))[0]
            second_folder = (item_values[0].split(": "))[1]
            log_path = os.path.join(folder_path, first_folder, second_folder, "log.txt")
            with open(log_path, "r") as file:
                lines = file.readlines()
            log_list = []

            for line in lines:
                if line.startswith("V"):
                    log_string = (
                        (line.split(":")[0])
                        + f": Enviado por {(((line.split(':')[1]).split(',')[0]).strip())}, {(((line.split(':')[1]).split(',')[1][1:11]))}"
                    )
                    log_list.append(log_string)
                if line.startswith("_"):
                    log_string = (line.split(":")[0])[
                        1:
                    ] + f": Revisto por {(((line.split(':')[1]).split(',')[0]).strip())}, {(((line.split(':')[1]).split(',')[1][1:11]))}"
                    log_list.append(log_string)
                if line.startswith("."):
                    log_string = (line.split(":")[0])[
                        1:
                    ] + f": Arquivado por {(((line.split(':')[1]).split(',')[0]).strip())}, {(((line.split(':')[1]).split(',')[1][1:11]))}"
                    log_list.append(log_string)
            filtered_log_list = []

            for item in log_list:
                
                if item not in filtered_log_list:
                    
                    filtered_log_list.append(item)

            
            self.print_box["state"] = "normal"
            self.print_box.delete(1.0, tk.END)
            for item in reversed(filtered_log_list):
                self.print_box.insert(tk.END, f"{item}\n")
            self.print_box["state"] = "disabled"

    def on_treeview_d_click(self, event):
        """Changes options.txt with the path to the project to be reviewed
        Also starts up the executable that to review the project
        """
        selected_item = self.treeview.selection()
        if selected_item:
            item_values = self.treeview.item(selected_item, "values")
            folder_item = item_values[0]
            parent_folder = (folder_item.split(":"))[0]
            target_folder = (folder_item.split(": "))[1]
            target_path = os.path.join(folder_path, parent_folder, target_folder)
            with open(options, "r") as file:
                lines = file.readlines()
            if lines:
                lines[0] = lines[0]
                lines[1] = f"path = {target_path}\n"
                lines[2] = f"phc = {folder_path_strip}\n"
            with open(options, "w") as file:
                file.writelines(lines)
        subprocess.Popen(compras_py, shell=True)


if __name__ == "__main__":
    dir_path = os.path.dirname(os.path.realpath(__file__))
    root = tk.Tk()
    root.tk.call("source", os.path.join(dir_path, "azure.tcl"))
    root.tk.call("set_theme", "light")
    app = FolderTreeViewApp(root)
    app.run()
